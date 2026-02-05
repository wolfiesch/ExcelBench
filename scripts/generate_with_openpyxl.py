#!/usr/bin/env python3
"""Alternative test file generator using openpyxl instead of xlwings.

This doesn't require Excel to be installed, but the generated files
might have minor differences from what Excel would produce.

Use this for development/testing when xlwings automation isn't available.
"""

import json
from datetime import UTC, date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUTPUT_DIR = Path("test_files")
GENERATOR_VERSION = "0.1.0-openpyxl"


def setup_header(ws):
    """Set up the header row."""
    ws["A1"] = "Label"
    ws["B1"] = "Test Cell"
    ws["C1"] = "Expected"

    for cell in ["A1", "B1", "C1"]:
        ws[cell].font = Font(bold=True)
        ws[cell].fill = PatternFill(start_color="FFDCDCDC", end_color="FFDCDCDC", fill_type="solid")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 50


def generate_cell_values():
    """Generate cell values test file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "cell_values"
    setup_header(ws)

    test_cases = []
    row = 2

    # String tests
    cases = [
        (
            "string_simple",
            "String - simple",
            "Hello World",
            {"type": "string", "value": "Hello World"},
        ),
        (
            "string_unicode",
            "String - unicode",
            "日本語🎉émojis",
            {"type": "string", "value": "日本語🎉émojis"},
        ),
        # Note: Excel converts empty strings to blank cells on save - this is correct Excel behavior
        ("string_empty", "String - empty", "", {"type": "blank"}),
        (
            "string_long",
            "String - long (1000 chars)",
            "A" * 1000,
            {"type": "string", "value": "A" * 1000},
        ),
        (
            "string_newline",
            "String - with newlines",
            "Line 1\nLine 2\nLine 3",
            {"type": "string", "value": "Line 1\nLine 2\nLine 3"},
        ),
        # Numbers
        ("number_integer", "Number - integer", 42, {"type": "number", "value": 42}),
        (
            "number_float",
            "Number - float",
            3.14159265358979,
            {"type": "number", "value": 3.14159265358979},
        ),
        (
            "number_negative",
            "Number - negative",
            -100.5,
            {"type": "number", "value": -100.5},
        ),
        (
            "number_large",
            "Number - large",
            1234567890123456,
            {"type": "number", "value": 1234567890123456},
        ),
        (
            "number_scientific",
            "Number - scientific notation",
            1.23e-10,
            {"type": "number", "value": 1.23e-10},
        ),
        # Booleans
        ("boolean_true", "Boolean - TRUE", True, {"type": "boolean", "value": True}),
        ("boolean_false", "Boolean - FALSE", False, {"type": "boolean", "value": False}),
    ]

    for case_id, label, value, expected in cases:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"C{row}"] = json.dumps(expected)
        test_cases.append({"id": case_id, "label": label, "row": row, "expected": expected})
        row += 1

    # Date
    ws[f"A{row}"] = "Date - standard"
    ws[f"B{row}"] = date(2026, 2, 4)
    ws[f"B{row}"].number_format = "yyyy-mm-dd"
    expected = {"type": "date", "value": "2026-02-04"}
    ws[f"C{row}"] = json.dumps(expected)
    test_cases.append({
        "id": "date_standard",
        "label": "Date - standard",
        "row": row,
        "expected": expected,
    })
    row += 1

    # DateTime
    ws[f"A{row}"] = "DateTime - with time"
    ws[f"B{row}"] = datetime(2026, 2, 4, 10, 30, 45)
    ws[f"B{row}"].number_format = "yyyy-mm-dd hh:mm:ss"
    expected = {"type": "datetime", "value": "2026-02-04T10:30:45"}
    ws[f"C{row}"] = json.dumps(expected)
    test_cases.append({
        "id": "datetime",
        "label": "DateTime - with time",
        "row": row,
        "expected": expected,
    })
    row += 1

    # Errors (via formulas)
    error_cases = [
        ("error_div0", "Error - #DIV/0!", "=1/0", {"type": "error", "value": "#DIV/0!"}),
        ("error_na", "Error - #N/A", "=NA()", {"type": "error", "value": "#N/A"}),
        ("error_value", "Error - #VALUE!", '="text"+1', {"type": "error", "value": "#VALUE!"}),
    ]

    for case_id, label, formula, expected in error_cases:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = formula
        ws[f"C{row}"] = json.dumps(expected)
        test_cases.append({"id": case_id, "label": label, "row": row, "expected": expected})
        row += 1

    # Blank
    ws[f"A{row}"] = "Blank cell"
    expected = {"type": "blank"}
    ws[f"C{row}"] = json.dumps(expected)
    test_cases.append({"id": "blank", "label": "Blank cell", "row": row, "expected": expected})

    # Save
    output_path = OUTPUT_DIR / "tier1" / "01_cell_values.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"  Created {output_path} with {len(test_cases)} test cases")

    return {
        "path": "tier1/01_cell_values.xlsx",
        "feature": "cell_values",
        "tier": 1,
        "test_cases": test_cases,
    }


def generate_text_formatting():
    """Generate text formatting test file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "text_formatting"
    setup_header(ws)

    test_cases = []
    row = 2

    cases = [
        ("bold", "Bold", Font(bold=True), {"bold": True}),
        ("italic", "Italic", Font(italic=True), {"italic": True}),
        (
            "underline_single",
            "Underline - single",
            Font(underline="single"),
            {"underline": "single"},
        ),
        (
            "underline_double",
            "Underline - double",
            Font(underline="double"),
            {"underline": "double"},
        ),
        ("strikethrough", "Strikethrough", Font(strike=True), {"strikethrough": True}),
        (
            "bold_italic",
            "Bold + Italic",
            Font(bold=True, italic=True),
            {"bold": True, "italic": True},
        ),
        ("font_size_8", "Font size 8", Font(size=8), {"font_size": 8}),
        ("font_size_14", "Font size 14", Font(size=14), {"font_size": 14}),
        ("font_size_24", "Font size 24", Font(size=24), {"font_size": 24}),
        ("font_size_36", "Font size 36", Font(size=36), {"font_size": 36}),
        ("font_arial", "Font - Arial", Font(name="Arial"), {"font_name": "Arial"}),
        (
            "font_times",
            "Font - Times New Roman",
            Font(name="Times New Roman"),
            {"font_name": "Times New Roman"},
        ),
        (
            "font_courier",
            "Font - Courier New",
            Font(name="Courier New"),
            {"font_name": "Courier New"},
        ),
        ("color_red", "Font color - red", Font(color="FF0000"), {"font_color": "#FF0000"}),
        ("color_blue", "Font color - blue", Font(color="0000FF"), {"font_color": "#0000FF"}),
        ("color_green", "Font color - green", Font(color="00FF00"), {"font_color": "#00FF00"}),
        (
            "color_custom",
            "Font color - custom (#8B4513)",
            Font(color="8B4513"),
            {"font_color": "#8B4513"},
        ),
        (
            "combined",
            "Combined - bold, 16pt, red",
            Font(bold=True, size=16, color="FF0000"),
            {"bold": True, "font_size": 16, "font_color": "#FF0000"},
        ),
    ]

    for case_id, label, font, expected in cases:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = f"{label} Text"
        ws[f"B{row}"].font = font
        ws[f"C{row}"] = json.dumps(expected)
        test_cases.append({"id": case_id, "label": label, "row": row, "expected": expected})
        row += 1

    output_path = OUTPUT_DIR / "tier1" / "03_text_formatting.xlsx"
    wb.save(output_path)
    print(f"  Created {output_path} with {len(test_cases)} test cases")

    return {
        "path": "tier1/03_text_formatting.xlsx",
        "feature": "text_formatting",
        "tier": 1,
        "test_cases": test_cases,
    }


def generate_formulas():
    """Generate formulas test file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "formulas"
    setup_header(ws)

    test_cases = []
    row = 2

    cases = [
        (
            "formula_sum",
            "Formula - SUM",
            "=SUM(1,2,3)",
            {"type": "formula", "formula": "=SUM(1,2,3)"},
        ),
        (
            "formula_cell_ref",
            "Formula - cell reference",
            "=A2*2",
            {"type": "formula", "formula": "=A2*2"},
        ),
        (
            "formula_concat",
            "Formula - concat",
            '=A2&" "&A3',
            {"type": "formula", "formula": '=A2&" "&A3'},
        ),
    ]

    # Seed referenced cells
    ws["A2"] = 10
    ws["A3"] = "World"

    for case_id, label, formula, expected in cases:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = formula
        ws[f"C{row}"] = json.dumps(expected)
        test_cases.append({"id": case_id, "label": label, "row": row, "expected": expected})
        row += 1

    # Cross-sheet reference
    ref = wb.create_sheet("References")
    setup_header(ref)
    ref["A2"] = "Reference"
    ref["B2"] = 42
    ref["C2"] = json.dumps({"type": "number", "value": 42})

    formula = "='References'!B2"
    expected = {"type": "formula", "formula": formula}
    ws[f"A{row}"] = "Formula - cross sheet"
    ws[f"B{row}"] = formula
    ws[f"C{row}"] = json.dumps(expected)
    test_cases.append({
        "id": "formula_cross_sheet",
        "label": "Formula - cross sheet",
        "row": row,
        "expected": expected,
    })

    output_path = OUTPUT_DIR / "tier1" / "02_formulas.xlsx"
    wb.save(output_path)
    print(f"  Created {output_path} with {len(test_cases)} test cases")

    return {
        "path": "tier1/02_formulas.xlsx",
        "feature": "formulas",
        "tier": 1,
        "test_cases": test_cases,
    }


def generate_background_colors():
    """Generate background color test file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "background_colors"
    setup_header(ws)

    test_cases = []
    row = 2

    cases = [
        ("bg_red", "Background - red", "FF0000", {"bg_color": "#FF0000"}),
        ("bg_blue", "Background - blue", "0000FF", {"bg_color": "#0000FF"}),
        ("bg_green", "Background - green", "00FF00", {"bg_color": "#00FF00"}),
        ("bg_custom", "Background - custom (#8B4513)", "8B4513", {"bg_color": "#8B4513"}),
    ]

    for case_id, label, color, expected in cases:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = label
        ws[f"B{row}"].fill = PatternFill(
            start_color=f"FF{color}",
            end_color=f"FF{color}",
            fill_type="solid",
        )
        ws[f"C{row}"] = json.dumps(expected)
        test_cases.append({"id": case_id, "label": label, "row": row, "expected": expected})
        row += 1

    output_path = OUTPUT_DIR / "tier1" / "04_background_colors.xlsx"
    wb.save(output_path)
    print(f"  Created {output_path} with {len(test_cases)} test cases")

    return {
        "path": "tier1/04_background_colors.xlsx",
        "feature": "background_colors",
        "tier": 1,
        "test_cases": test_cases,
    }


def generate_number_formats():
    """Generate number formats test file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "number_formats"
    setup_header(ws)

    test_cases = []
    row = 2

    cases = [
        (
            "numfmt_currency",
            "Format - currency",
            1234.56,
            "$#,##0.00",
            {"number_format": "$#,##0.00"},
        ),
        (
            "numfmt_percent",
            "Format - percent",
            0.256,
            "0.00%",
            {"number_format": "0.00%"},
        ),
        (
            "numfmt_date",
            "Format - date",
            date(2026, 2, 4),
            "yyyy-mm-dd",
            {"number_format": "yyyy-mm-dd"},
        ),
        (
            "numfmt_scientific",
            "Format - scientific",
            12345.678,
            "0.00E+00",
            {"number_format": "0.00E+00"},
        ),
        (
            "numfmt_custom",
            "Format - custom text",
            12.3,
            "\"USD\" 0.00",
            {"number_format": "\"USD\" 0.00"},
        ),
    ]

    for case_id, label, value, fmt, expected in cases:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        ws[f"B{row}"].number_format = fmt
        ws[f"C{row}"] = json.dumps(expected)
        test_cases.append({"id": case_id, "label": label, "row": row, "expected": expected})
        row += 1

    output_path = OUTPUT_DIR / "tier1" / "05_number_formats.xlsx"
    wb.save(output_path)
    print(f"  Created {output_path} with {len(test_cases)} test cases")

    return {
        "path": "tier1/05_number_formats.xlsx",
        "feature": "number_formats",
        "tier": 1,
        "test_cases": test_cases,
    }


def generate_alignment():
    """Generate alignment test file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "alignment"
    setup_header(ws)

    test_cases = []
    row = 2

    cases = [
        ("h_left", "Align - left", Alignment(horizontal="left"), {"h_align": "left"}),
        ("h_center", "Align - center", Alignment(horizontal="center"), {"h_align": "center"}),
        ("h_right", "Align - right", Alignment(horizontal="right"), {"h_align": "right"}),
        ("v_top", "Align - top", Alignment(vertical="top"), {"v_align": "top"}),
        ("v_center", "Align - center", Alignment(vertical="center"), {"v_align": "center"}),
        ("v_bottom", "Align - bottom", Alignment(vertical="bottom"), {"v_align": "bottom"}),
        ("wrap", "Align - wrap text", Alignment(wrap_text=True), {"wrap": True}),
        ("rotation", "Align - rotation 45", Alignment(text_rotation=45), {"rotation": 45}),
        ("indent", "Align - indent 2", Alignment(indent=2), {"indent": 2}),
    ]

    for case_id, label, align, expected in cases:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = label
        ws[f"B{row}"].alignment = align
        ws[f"C{row}"] = json.dumps(expected)
        test_cases.append({"id": case_id, "label": label, "row": row, "expected": expected})
        row += 1

    output_path = OUTPUT_DIR / "tier1" / "06_alignment.xlsx"
    wb.save(output_path)
    print(f"  Created {output_path} with {len(test_cases)} test cases")

    return {
        "path": "tier1/06_alignment.xlsx",
        "feature": "alignment",
        "tier": 1,
        "test_cases": test_cases,
    }


def generate_dimensions():
    """Generate dimensions test file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "dimensions"
    setup_header(ws)

    test_cases = []
    row = 2

    # Row height tests
    for height, case_id in [(30, "row_height_30"), (45, "row_height_45")]:
        label = f"Row height - {height}"
        expected = {"row_height": height}
        ws[f"A{row}"] = label
        ws[f"B{row}"] = label
        ws[f"C{row}"] = json.dumps(expected)
        ws.row_dimensions[row].height = height
        test_cases.append({
            "id": case_id,
            "label": label,
            "row": row,
            "expected": expected,
            "cell": f"B{row}",
        })
        row += 1

    # Column width tests (columns D/E)
    for column, width, case_id in [("D", 20, "col_width_20"), ("E", 8, "col_width_8")]:
        label = f"Column width - {column} = {width}"
        expected = {"column_width": width}
        ws[f"A{row}"] = label
        ws[f"{column}{row}"] = label
        ws[f"C{row}"] = json.dumps(expected)
        ws.column_dimensions[column].width = width
        test_cases.append({
            "id": case_id,
            "label": label,
            "row": row,
            "expected": expected,
            "cell": f"{column}{row}",
        })
        row += 1

    output_path = OUTPUT_DIR / "tier1" / "08_dimensions.xlsx"
    wb.save(output_path)
    print(f"  Created {output_path} with {len(test_cases)} test cases")

    return {
        "path": "tier1/08_dimensions.xlsx",
        "feature": "dimensions",
        "tier": 1,
        "test_cases": test_cases,
    }


def generate_multiple_sheets():
    """Generate multiple sheets test file."""
    wb = Workbook()
    ws_alpha = wb.active
    ws_alpha.title = "Alpha"
    setup_header(ws_alpha)

    ws_beta = wb.create_sheet("Beta")
    setup_header(ws_beta)
    ws_gamma = wb.create_sheet("Gamma")
    setup_header(ws_gamma)

    test_cases = []

    expected = {"sheet_names": ["Alpha", "Beta", "Gamma"]}
    ws_alpha["A2"] = "Sheet names"
    ws_alpha["C2"] = json.dumps(expected)
    test_cases.append({
        "id": "sheet_names",
        "label": "Sheet names",
        "row": 2,
        "expected": expected,
        "sheet": "Alpha",
    })

    for ws, name in [(ws_alpha, "Alpha"), (ws_beta, "Beta"), (ws_gamma, "Gamma")]:
        label = f"{name} value"
        expected = {"type": "string", "value": name}
        ws["A3"] = label
        ws["B3"] = name
        ws["C3"] = json.dumps(expected)
        test_cases.append({
            "id": f"value_{name.lower()}",
            "label": label,
            "row": 3,
            "expected": expected,
            "sheet": name,
            "cell": "B3",
        })

    output_path = OUTPUT_DIR / "tier1" / "09_multiple_sheets.xlsx"
    wb.save(output_path)
    print(f"  Created {output_path} with {len(test_cases)} test cases")

    return {
        "path": "tier1/09_multiple_sheets.xlsx",
        "feature": "multiple_sheets",
        "tier": 1,
        "test_cases": test_cases,
    }


def generate_borders():
    """Generate borders test file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "borders"
    setup_header(ws)

    test_cases = []
    row = 2

    def make_border(style="thin", color="000000"):
        side = Side(style=style, color=color)
        return Border(left=side, right=side, top=side, bottom=side)

    cases = [
        (
            "thin_all",
            "Border - thin all edges",
            make_border("thin"),
            {"border_style": "thin", "border_color": "#000000"},
        ),
        (
            "medium_all",
            "Border - medium all edges",
            make_border("medium"),
            {"border_style": "medium", "border_color": "#000000"},
        ),
        (
            "thick_all",
            "Border - thick all edges",
            make_border("thick"),
            {"border_style": "thick", "border_color": "#000000"},
        ),
        (
            "double",
            "Border - double line",
            make_border("double"),
            {"border_style": "double", "border_color": "#000000"},
        ),
        (
            "dashed",
            "Border - dashed",
            make_border("dashed"),
            {"border_style": "dashed", "border_color": "#000000"},
        ),
        (
            "dotted",
            "Border - dotted",
            make_border("dotted"),
            {"border_style": "dotted", "border_color": "#000000"},
        ),
        (
            "dash_dot",
            "Border - dash-dot",
            make_border("dashDot"),
            {"border_style": "dashDot", "border_color": "#000000"},
        ),
        (
            "dash_dot_dot",
            "Border - dash-dot-dot",
            make_border("dashDotDot"),
            {"border_style": "dashDotDot", "border_color": "#000000"},
        ),
    ]

    for case_id, label, border, expected in cases:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = label.split(" - ")[1] if " - " in label else label
        ws[f"B{row}"].border = border
        ws[f"C{row}"] = json.dumps(expected)
        test_cases.append({"id": case_id, "label": label, "row": row, "expected": expected})
        row += 1

    # Individual edges
    edge_cases = [
        (
            "top_only",
            "Border - top only",
            Border(top=Side(style="thin")),
            {
                "border_top": "thin",
                "border_bottom": None,
                "border_left": None,
                "border_right": None,
            },
        ),
        (
            "bottom_only",
            "Border - bottom only",
            Border(bottom=Side(style="thin")),
            {
                "border_top": None,
                "border_bottom": "thin",
                "border_left": None,
                "border_right": None,
            },
        ),
        (
            "left_only",
            "Border - left only",
            Border(left=Side(style="thin")),
            {
                "border_top": None,
                "border_bottom": None,
                "border_left": "thin",
                "border_right": None,
            },
        ),
        (
            "right_only",
            "Border - right only",
            Border(right=Side(style="thin")),
            {
                "border_top": None,
                "border_bottom": None,
                "border_left": None,
                "border_right": "thin",
            },
        ),
    ]

    for case_id, label, border, expected in edge_cases:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = label.split(" - ")[1]
        ws[f"B{row}"].border = border
        ws[f"C{row}"] = json.dumps(expected)
        test_cases.append({"id": case_id, "label": label, "row": row, "expected": expected})
        row += 1

    # Diagonal borders
    ws[f"A{row}"] = "Border - diagonal up"
    ws[f"B{row}"] = "Diag Up"
    ws[f"B{row}"].border = Border(diagonal=Side(style="thin"), diagonalUp=True)
    expected = {"border_diagonal_up": "thin"}
    ws[f"C{row}"] = json.dumps(expected)
    test_cases.append(
        {
            "id": "diagonal_up",
            "label": "Border - diagonal up",
            "row": row,
            "expected": expected,
        }
    )
    row += 1

    ws[f"A{row}"] = "Border - diagonal down"
    ws[f"B{row}"] = "Diag Down"
    ws[f"B{row}"].border = Border(diagonal=Side(style="thin"), diagonalDown=True)
    expected = {"border_diagonal_down": "thin"}
    ws[f"C{row}"] = json.dumps(expected)
    test_cases.append(
        {
            "id": "diagonal_down",
            "label": "Border - diagonal down",
            "row": row,
            "expected": expected,
        }
    )
    row += 1

    # Colors
    color_cases = [
        (
            "color_red",
            "Border - red color",
            make_border(color="FF0000"),
            {"border_style": "thin", "border_color": "#FF0000"},
        ),
        (
            "color_blue",
            "Border - blue color",
            make_border(color="0000FF"),
            {"border_style": "thin", "border_color": "#0000FF"},
        ),
        (
            "color_custom",
            "Border - custom color (#8B4513)",
            make_border(color="8B4513"),
            {"border_style": "thin", "border_color": "#8B4513"},
        ),
    ]

    for case_id, label, border, expected in color_cases:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = label.split(" - ")[1]
        ws[f"B{row}"].border = border
        ws[f"C{row}"] = json.dumps(expected)
        test_cases.append({"id": case_id, "label": label, "row": row, "expected": expected})
        row += 1

    output_path = OUTPUT_DIR / "tier1" / "07_borders.xlsx"
    wb.save(output_path)
    print(f"  Created {output_path} with {len(test_cases)} test cases")

    return {
        "path": "tier1/07_borders.xlsx",
        "feature": "borders",
        "tier": 1,
        "test_cases": test_cases,
    }


def main():
    print("Generating test files with openpyxl...")
    print()

    files = []
    files.append(generate_cell_values())
    files.append(generate_formulas())
    files.append(generate_text_formatting())
    files.append(generate_background_colors())
    files.append(generate_number_formats())
    files.append(generate_alignment())
    files.append(generate_borders())
    files.append(generate_dimensions())
    files.append(generate_multiple_sheets())

    # Write manifest
    manifest = {
        "generated_at": datetime.now(UTC).isoformat(),
        "excel_version": "openpyxl-generated",
        "generator_version": GENERATOR_VERSION,
        "files": files,
    }

    manifest_path = OUTPUT_DIR / "manifest.json"
    with open(manifest_path, "w") as f:
        json.dump(manifest, f, indent=2)

    print()
    print(f"✓ Generated {len(files)} test files")
    print(f"  Manifest: {manifest_path}")


if __name__ == "__main__":
    main()
