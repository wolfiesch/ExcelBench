#!/usr/bin/env python3
"""Generate throughput/scale performance fixtures.

These fixtures are intended for *performance* benchmarking (excelbench perf), not fidelity.
They use a compact workload spec in `expected.workload` to avoid huge manifests.

Default output is under `test_files/` so it stays gitignored.
"""

from __future__ import annotations

import argparse
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side

from excelbench.generator.generate import write_manifest
from excelbench.models import Importance, Manifest, TestCase, TestFile


def _coord_to_cell(row: int, col: int) -> str:
    letters = ""
    c = col
    while c > 0:
        c, rem = divmod(c - 1, 26)
        letters = chr(65 + rem) + letters
    return f"{letters}{row}"


def _generate_cell_values_grid(
    *,
    path: Path,
    sheet: str,
    rows: int,
    cols: int,
    start: int = 1,
    step: int = 1,
) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet

    value = start
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c, value=value)
            value += step

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _generate_formulas_grid(
    *,
    path: Path,
    sheet: str,
    rows: int,
    cols: int,
    formula: str = "=1+1",
) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet

    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c, value=formula)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _generate_bg_colors_grid(
    *,
    path: Path,
    sheet: str,
    rows: int,
    cols: int,
    palette: list[str],
) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet

    fills = [
        PatternFill(start_color=f"FF{c}", end_color=f"FF{c}", fill_type="solid") for c in palette
    ]

    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c, value="Color")
            cell.fill = fills[(r * cols + c) % len(fills)]

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _generate_number_formats_grid(
    *,
    path: Path,
    sheet: str,
    rows: int,
    cols: int,
    number_format: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet

    value = 0.5
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.number_format = number_format
            value += 1.0

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _generate_alignment_grid(
    *,
    path: Path,
    sheet: str,
    rows: int,
    cols: int,
    alignment: Alignment,
) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet

    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c, value="Align")
            cell.alignment = alignment

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _generate_borders_grid(
    *,
    path: Path,
    sheet: str,
    rows: int,
    cols: int,
    border: Border,
) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet

    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c, value="Border")
            cell.border = border

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate ExcelBench throughput fixtures")
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        default=Path("test_files/throughput_xlsx"),
        help="Output directory (default: test_files/throughput_xlsx)",
    )
    parser.add_argument(
        "--include-100k",
        action="store_true",
        help="Also generate a ~100k-cell fixture (can take a while).",
    )
    args = parser.parse_args()

    out = Path(args.output)
    tier_dir = out / "tier0"
    tier_dir.mkdir(parents=True, exist_ok=True)

    files: list[TestFile] = []

    # 10k = 100x100
    scenario = "cell_values_10k"
    sheet = "S1"
    rows, cols = 100, 100
    end_cell = _coord_to_cell(rows, cols)
    rng = f"A1:{end_cell}"
    filename = "00_cell_values_10k.xlsx"
    _generate_cell_values_grid(path=tier_dir / filename, sheet=sheet, rows=rows, cols=cols)
    files.append(
        TestFile(
            path=f"tier0/{filename}",
            feature=scenario,
            tier=0,
            file_format="xlsx",
            test_cases=[
                TestCase(
                    id=scenario,
                    label="Throughput: cell values (10k cells)",
                    row=1,
                    expected={
                        "workload": {
                            "scenario": scenario,
                            "op": "cell_value",
                            "sheet": sheet,
                            "range": rng,
                            "start": 1,
                            "step": 1,
                        }
                    },
                    importance=Importance.BASIC,
                )
            ],
        )
    )

    # 1k = 40x25 (useful for very slow per-cell readers)
    scenario = "cell_values_1k"
    sheet = "S1"
    rows, cols = 40, 25
    end_cell = _coord_to_cell(rows, cols)
    rng = f"A1:{end_cell}"
    filename = "00_cell_values_1k.xlsx"
    _generate_cell_values_grid(path=tier_dir / filename, sheet=sheet, rows=rows, cols=cols)
    files.append(
        TestFile(
            path=f"tier0/{filename}",
            feature=scenario,
            tier=0,
            file_format="xlsx",
            test_cases=[
                TestCase(
                    id=scenario,
                    label="Throughput: cell values (1k cells)",
                    row=1,
                    expected={
                        "workload": {
                            "scenario": scenario,
                            "op": "cell_value",
                            "sheet": sheet,
                            "range": rng,
                            "start": 1,
                            "step": 1,
                        }
                    },
                    importance=Importance.BASIC,
                )
            ],
        )
    )

    # 10k formulas = 100x100
    scenario = "formulas_10k"
    sheet = "S1"
    rows, cols = 100, 100
    end_cell = _coord_to_cell(rows, cols)
    rng = f"A1:{end_cell}"
    filename = "00_formulas_10k.xlsx"
    formula = "=1+1"
    _generate_formulas_grid(
        path=tier_dir / filename,
        sheet=sheet,
        rows=rows,
        cols=cols,
        formula=formula,
    )
    files.append(
        TestFile(
            path=f"tier0/{filename}",
            feature=scenario,
            tier=0,
            file_format="xlsx",
            test_cases=[
                TestCase(
                    id=scenario,
                    label="Throughput: formulas (10k cells)",
                    row=1,
                    expected={
                        "workload": {
                            "scenario": scenario,
                            "op": "formula",
                            "sheet": sheet,
                            "range": rng,
                            "formula": formula,
                        }
                    },
                    importance=Importance.BASIC,
                )
            ],
        )
    )

    # 1k formulas = 40x25
    scenario = "formulas_1k"
    sheet = "S1"
    rows, cols = 40, 25
    end_cell = _coord_to_cell(rows, cols)
    rng = f"A1:{end_cell}"
    filename = "00_formulas_1k.xlsx"
    formula = "=1+1"
    _generate_formulas_grid(
        path=tier_dir / filename,
        sheet=sheet,
        rows=rows,
        cols=cols,
        formula=formula,
    )
    files.append(
        TestFile(
            path=f"tier0/{filename}",
            feature=scenario,
            tier=0,
            file_format="xlsx",
            test_cases=[
                TestCase(
                    id=scenario,
                    label="Throughput: formulas (1k cells)",
                    row=1,
                    expected={
                        "workload": {
                            "scenario": scenario,
                            "op": "formula",
                            "sheet": sheet,
                            "range": rng,
                            "formula": formula,
                        }
                    },
                    importance=Importance.BASIC,
                )
            ],
        )
    )

    # 1k background fills = 40x25
    scenario = "background_colors_1k"
    sheet = "S1"
    rows, cols = 40, 25
    end_cell = _coord_to_cell(rows, cols)
    rng = f"A1:{end_cell}"
    filename = "00_background_colors_1k.xlsx"
    palette = ["FF0000", "00FF00", "0000FF", "FFFF00"]
    _generate_bg_colors_grid(
        path=tier_dir / filename,
        sheet=sheet,
        rows=rows,
        cols=cols,
        palette=palette,
    )
    files.append(
        TestFile(
            path=f"tier0/{filename}",
            feature=scenario,
            tier=0,
            file_format="xlsx",
            test_cases=[
                TestCase(
                    id=scenario,
                    label="Throughput: background fills (1k cells)",
                    row=1,
                    expected={
                        "workload": {
                            "scenario": scenario,
                            "op": "bg_color",
                            "sheet": sheet,
                            "range": rng,
                            "palette": [f"#{c}" for c in palette],
                        }
                    },
                    importance=Importance.BASIC,
                )
            ],
        )
    )

    # 1k number formats = 40x25
    scenario = "number_formats_1k"
    sheet = "S1"
    rows, cols = 40, 25
    end_cell = _coord_to_cell(rows, cols)
    rng = f"A1:{end_cell}"
    filename = "00_number_formats_1k.xlsx"
    number_format = "0.00%"
    _generate_number_formats_grid(
        path=tier_dir / filename,
        sheet=sheet,
        rows=rows,
        cols=cols,
        number_format=number_format,
    )
    files.append(
        TestFile(
            path=f"tier0/{filename}",
            feature=scenario,
            tier=0,
            file_format="xlsx",
            test_cases=[
                TestCase(
                    id=scenario,
                    label="Throughput: number formats (1k cells)",
                    row=1,
                    expected={
                        "workload": {
                            "scenario": scenario,
                            "op": "number_format",
                            "sheet": sheet,
                            "range": rng,
                            "number_format": number_format,
                        }
                    },
                    importance=Importance.BASIC,
                )
            ],
        )
    )

    # 1k alignment = 40x25
    scenario = "alignment_1k"
    sheet = "S1"
    rows, cols = 40, 25
    end_cell = _coord_to_cell(rows, cols)
    rng = f"A1:{end_cell}"
    filename = "00_alignment_1k.xlsx"
    align = Alignment(horizontal="center", vertical="top", wrap_text=True)
    _generate_alignment_grid(
        path=tier_dir / filename,
        sheet=sheet,
        rows=rows,
        cols=cols,
        alignment=align,
    )
    files.append(
        TestFile(
            path=f"tier0/{filename}",
            feature=scenario,
            tier=0,
            file_format="xlsx",
            test_cases=[
                TestCase(
                    id=scenario,
                    label="Throughput: alignment (1k cells)",
                    row=1,
                    expected={
                        "workload": {
                            "scenario": scenario,
                            "op": "alignment",
                            "sheet": sheet,
                            "range": rng,
                            "h_align": "center",
                            "v_align": "top",
                            "wrap": True,
                        }
                    },
                    importance=Importance.BASIC,
                )
            ],
        )
    )

    # 200 borders = 20x10
    scenario = "borders_200"
    sheet = "S1"
    rows, cols = 20, 10
    end_cell = _coord_to_cell(rows, cols)
    rng = f"A1:{end_cell}"
    filename = "00_borders_200.xlsx"
    side = Side(style="thin", color="FF000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    _generate_borders_grid(
        path=tier_dir / filename,
        sheet=sheet,
        rows=rows,
        cols=cols,
        border=border,
    )
    files.append(
        TestFile(
            path=f"tier0/{filename}",
            feature=scenario,
            tier=0,
            file_format="xlsx",
            test_cases=[
                TestCase(
                    id=scenario,
                    label="Throughput: borders (200 cells)",
                    row=1,
                    expected={
                        "workload": {
                            "scenario": scenario,
                            "op": "border",
                            "sheet": sheet,
                            "range": rng,
                            "border_style": "thin",
                            "border_color": "#000000",
                        }
                    },
                    importance=Importance.BASIC,
                )
            ],
        )
    )

    if args.include_100k:
        # ~100k = 316x316 = 99856 cells
        scenario = "cell_values_100k"
        sheet = "S1"
        rows, cols = 316, 316
        end_cell = _coord_to_cell(rows, cols)
        rng = f"A1:{end_cell}"
        filename = "00_cell_values_100k.xlsx"
        _generate_cell_values_grid(path=tier_dir / filename, sheet=sheet, rows=rows, cols=cols)
        files.append(
            TestFile(
                path=f"tier0/{filename}",
                feature=scenario,
                tier=0,
                file_format="xlsx",
                test_cases=[
                    TestCase(
                        id=scenario,
                        label="Throughput: cell values (~100k cells)",
                        row=1,
                        expected={
                            "workload": {
                                "scenario": scenario,
                                "op": "cell_value",
                                "sheet": sheet,
                                "range": rng,
                                "start": 1,
                                "step": 1,
                            }
                        },
                        importance=Importance.BASIC,
                    )
                ],
            )
        )

    manifest = Manifest(
        generated_at=datetime.now(UTC),
        excel_version="openpyxl-generated",
        generator_version="throughput-0.1.0",
        file_format="xlsx",
        files=files,
    )
    write_manifest(manifest, out / "manifest.json")

    print(f"✓ Wrote {len(files)} throughput fixture(s) to {out}")
    print(f"  Manifest: {out / 'manifest.json'}")


if __name__ == "__main__":
    main()
