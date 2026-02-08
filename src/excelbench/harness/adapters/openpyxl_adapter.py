"""Adapter for openpyxl library."""

from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles import colors as _openpyxl_colors
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink

from excelbench.harness.adapters.base import ExcelAdapter
from excelbench.models import (
    BorderEdge,
    BorderInfo,
    BorderStyle,
    CellFormat,
    CellType,
    CellValue,
    LibraryInfo,
)

JSONDict = dict[str, Any]

# Formulas that produce known error values (openpyxl returns formula, not error)
ERROR_FORMULA_MAP = {
    "=1/0": "#DIV/0!",
    "=NA()": "#N/A",
    '="text"+1': "#VALUE!",
}


_COLOR_INDEX = getattr(_openpyxl_colors, "COLOR_INDEX", None)


def _openpyxl_color_to_hex(color: Any) -> str | None:
    if not color:
        return None

    rgb = getattr(color, "rgb", None)
    rgb_str: str | None = None
    if isinstance(rgb, str):
        rgb_str = rgb
    elif isinstance(rgb, (bytes, bytearray)):
        try:
            rgb_str = bytes(rgb).decode("ascii", errors="ignore")
        except Exception:
            rgb_str = None
    else:
        value = getattr(rgb, "value", None)
        if isinstance(value, str):
            rgb_str = value

    if isinstance(rgb_str, str) and len(rgb_str) >= 6:
        if len(rgb_str) == 8:
            return f"#{rgb_str[2:]}"  # Skip alpha (ARGB)
        return f"#{rgb_str}"

    indexed = getattr(color, "indexed", None)
    if isinstance(indexed, int) and _COLOR_INDEX is not None:
        try:
            argb = _COLOR_INDEX[indexed]
            if isinstance(argb, str) and len(argb) == 8:
                return f"#{argb[2:]}"
        except Exception:
            return None

    return None


def _col_letter(index: int) -> str:
    result = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        result = chr(65 + rem) + result
    return result


def _get_version() -> str:
    """Get openpyxl version."""
    return str(openpyxl.__version__)


class OpenpyxlAdapter(ExcelAdapter):
    """Adapter for openpyxl library (read/write support)."""

    @property
    def info(self) -> LibraryInfo:
        return LibraryInfo(
            name="openpyxl",
            version=_get_version(),
            language="python",
            capabilities={"read", "write"},
        )

    @property
    def supported_read_extensions(self) -> set[str]:
        return {".xlsx"}

    # =========================================================================
    # Read Operations
    # =========================================================================

    def open_workbook(self, path: Path) -> Workbook:
        """Open a workbook for reading."""
        return openpyxl.load_workbook(str(path), data_only=False)

    def close_workbook(self, workbook: Any) -> None:
        """Close an opened workbook."""
        workbook.close()

    def get_sheet_names(self, workbook: Workbook) -> list[str]:
        """Get list of sheet names in a workbook."""
        return [str(name) for name in workbook.sheetnames]

    def read_cell_value(
        self,
        workbook: Workbook,
        sheet: str,
        cell: str,
    ) -> CellValue:
        """Read the value of a cell."""
        ws = workbook[sheet]
        c: Cell = ws[cell]

        # Handle different value types
        value = c.value

        if value is None:
            return CellValue(type=CellType.BLANK)

        if isinstance(value, bool):
            return CellValue(type=CellType.BOOLEAN, value=value)

        if isinstance(value, (int, float)):
            return CellValue(type=CellType.NUMBER, value=value)

        # Check date BEFORE datetime since datetime is a subclass of date
        if isinstance(value, date) and not isinstance(value, datetime):
            return CellValue(type=CellType.DATE, value=value)

        if isinstance(value, datetime):
            # Check if this is a "date" (time component is midnight)
            # Excel stores dates as datetimes with 00:00:00 time
            if (
                value.hour == 0
                and value.minute == 0
                and value.second == 0
                and value.microsecond == 0
            ):
                return CellValue(type=CellType.DATE, value=value.date())
            return CellValue(type=CellType.DATETIME, value=value)

        if isinstance(value, str):
            # Check if it's an error value
            if value in ("#N/A", "#NULL!", "#NAME?", "#REF!"):
                return CellValue(type=CellType.ERROR, value=value)
            if value.startswith("#") and value.endswith("!"):
                return CellValue(type=CellType.ERROR, value=value)

            # Check if there's a formula
            if c.data_type == "f" or (hasattr(c, "value") and str(c.value).startswith("=")):
                # Check if this formula produces a known error value
                formula_str = str(c.value)
                if formula_str and not formula_str.startswith("="):
                    formula_str = f"={formula_str}"
                if not formula_str.startswith("=") and value:
                    formula_str = str(value)
                if formula_str in ERROR_FORMULA_MAP:
                    return CellValue(type=CellType.ERROR, value=ERROR_FORMULA_MAP[formula_str])
                return CellValue(
                    type=CellType.FORMULA,
                    value=value,
                    formula=formula_str,
                )

            return CellValue(type=CellType.STRING, value=value)

        # Fallback to string
        return CellValue(type=CellType.STRING, value=str(value))

    def read_cell_format(
        self,
        workbook: Workbook,
        sheet: str,
        cell: str,
    ) -> CellFormat:
        """Read the formatting of a cell."""
        ws = workbook[sheet]
        c: Cell = ws[cell]
        font = c.font

        # Convert color to hex
        font_color = _openpyxl_color_to_hex(getattr(font, "color", None))

        # Get background color
        bg_color = None
        fill = c.fill
        if fill and getattr(fill, "patternType", None) == "solid":
            bg_color = _openpyxl_color_to_hex(getattr(fill, "fgColor", None))

        # Map underline
        underline = None
        if font.underline:
            underline_map = {
                "single": "single",
                "double": "double",
                "singleAccounting": "singleAccounting",
                "doubleAccounting": "doubleAccounting",
            }
            underline = underline_map.get(font.underline, font.underline)

        alignment = c.alignment
        h_align = alignment.horizontal if alignment and alignment.horizontal else None
        v_align = alignment.vertical if alignment and alignment.vertical else None
        wrap = alignment.wrap_text if alignment and alignment.wrap_text else None
        rotation = (
            alignment.text_rotation
            if alignment and alignment.text_rotation not in (0, None)
            else None
        )
        indent = alignment.indent if alignment and alignment.indent else None

        return CellFormat(
            bold=font.bold if font.bold else None,
            italic=font.italic if font.italic else None,
            underline=underline,
            strikethrough=font.strike if font.strike else None,
            font_name=font.name if font.name else None,
            font_size=font.size if font.size else None,
            font_color=font_color,
            bg_color=bg_color,
            number_format=c.number_format if c.number_format else None,
            h_align=h_align,
            v_align=v_align,
            wrap=wrap,
            rotation=rotation,
            indent=indent,
        )

    def read_cell_border(
        self,
        workbook: Workbook,
        sheet: str,
        cell: str,
    ) -> BorderInfo:
        """Read the border information of a cell."""
        ws = workbook[sheet]
        c: Cell = ws[cell]
        border = c.border

        def parse_side(side: Side | None) -> BorderEdge | None:
            if side is None or side.style is None:
                return None

            # Map openpyxl style to our style
            style_map = {
                "thin": BorderStyle.THIN,
                "medium": BorderStyle.MEDIUM,
                "thick": BorderStyle.THICK,
                "double": BorderStyle.DOUBLE,
                "dashed": BorderStyle.DASHED,
                "dotted": BorderStyle.DOTTED,
                "hair": BorderStyle.HAIR,
                "mediumDashed": BorderStyle.MEDIUM_DASHED,
                "dashDot": BorderStyle.DASH_DOT,
                "mediumDashDot": BorderStyle.MEDIUM_DASH_DOT,
                "dashDotDot": BorderStyle.DASH_DOT_DOT,
                "mediumDashDotDot": BorderStyle.MEDIUM_DASH_DOT_DOT,
                "slantDashDot": BorderStyle.SLANT_DASH_DOT,
            }

            style = style_map.get(side.style, BorderStyle.THIN)

            # Get color
            color = _openpyxl_color_to_hex(getattr(side, "color", None)) or "#000000"

            return BorderEdge(style=style, color=color)

        return BorderInfo(
            top=parse_side(border.top),
            bottom=parse_side(border.bottom),
            left=parse_side(border.left),
            right=parse_side(border.right),
            diagonal_up=parse_side(border.diagonal) if border.diagonalUp else None,
            diagonal_down=parse_side(border.diagonal) if border.diagonalDown else None,
        )

    def read_row_height(
        self,
        workbook: Workbook,
        sheet: str,
        row: int,
    ) -> float | None:
        ws = workbook[sheet]
        height = ws.row_dimensions[row].height
        return float(height) if isinstance(height, (int, float)) else None

    def read_column_width(
        self,
        workbook: Workbook,
        sheet: str,
        column: str,
    ) -> float | None:
        ws = workbook[sheet]
        width = ws.column_dimensions[column].width
        if width is None:
            return None
        try:
            width_f = float(width)
        except (TypeError, ValueError):
            return None
        # Excel and third-party libraries add font-metric padding to stored
        # column widths. Known paddings:
        #   0.83203125 - Excel (Calibri 11pt default)
        #   0.7109375  - xlsxwriter
        # Strip the padding to return the display character width.
        frac = width_f % 1
        known_paddings = [0.83203125, 0.7109375]
        for padding in known_paddings:
            if abs(frac - padding) < 0.01:
                width_f = width_f - padding
                break
        return round(width_f, 4)

    # =========================================================================
    # Tier 2 Read Operations
    # =========================================================================

    def read_merged_ranges(self, workbook: Workbook, sheet: str) -> list[str]:
        ws = workbook[sheet]
        return [str(rng) for rng in ws.merged_cells.ranges]

    def read_conditional_formats(self, workbook: Workbook, sheet: str) -> list[JSONDict]:
        ws = workbook[sheet]
        rules: list[JSONDict] = []
        cf_rules = getattr(ws.conditional_formatting, "_cf_rules", {})
        for sqref, rule_list in cf_rules.items():
            range_value = None
            if hasattr(sqref, "sqref"):
                range_value = str(sqref.sqref)
            else:
                range_value = str(sqref)
                if range_value.startswith("<ConditionalFormatting"):
                    range_value = (
                        range_value.replace("<ConditionalFormatting", "").replace(">", "").strip()
                    )
            for rule in rule_list:
                entry: dict[str, Any] = {
                    "range": range_value,
                    "rule_type": getattr(rule, "type", None),
                    "operator": getattr(rule, "operator", None),
                    "formula": None,
                    "priority": getattr(rule, "priority", None),
                    "stop_if_true": getattr(rule, "stopIfTrue", None),
                    "format": {},
                }
                if getattr(rule, "formula", None):
                    entry["formula"] = rule.formula[0] if rule.formula else None
                dxf = getattr(rule, "dxf", None)
                if dxf and getattr(dxf, "fill", None) and getattr(dxf.fill, "fgColor", None):
                    bg = _openpyxl_color_to_hex(dxf.fill.fgColor)
                    if bg:
                        entry["format"]["bg_color"] = bg
                if dxf and getattr(dxf, "font", None) and getattr(dxf.font, "color", None):
                    fc = _openpyxl_color_to_hex(dxf.font.color)
                    if fc:
                        entry["format"]["font_color"] = fc
                rules.append(entry)
        return rules

    def read_data_validations(self, workbook: Workbook, sheet: str) -> list[JSONDict]:
        ws = workbook[sheet]
        validations: list[JSONDict] = []
        dv = getattr(ws, "data_validations", None)
        if not dv:
            return validations
        for entry in dv.dataValidation:
            # "between" is the xlsx default operator; libraries may omit it
            operator = entry.operator
            if operator is None and entry.formula2:
                operator = "between"
            validations.append(
                {
                    "range": str(entry.sqref),
                    "validation_type": entry.type,
                    "operator": operator,
                    "formula1": entry.formula1,
                    "formula2": entry.formula2,
                    "allow_blank": entry.allow_blank,
                    "show_input": entry.showInputMessage,
                    "show_error": entry.showErrorMessage,
                    "prompt_title": entry.promptTitle,
                    "prompt": entry.prompt,
                    "error_title": entry.errorTitle,
                    "error": entry.error,
                }
            )
        return validations

    def read_hyperlinks(self, workbook: Workbook, sheet: str) -> list[JSONDict]:
        ws = workbook[sheet]
        links: list[JSONDict] = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.hyperlink:
                    h = cell.hyperlink
                    # xlsx stores URL and fragment separately (target + location).
                    # Recombine when both exist (e.g. https://...#section-2).
                    if h.target and h.location:
                        target = f"{h.target}#{h.location}"
                        internal = False
                    elif h.target:
                        target = h.target
                        internal = False
                    else:
                        target = h.location
                        internal = True
                    links.append(
                        {
                            "cell": cell.coordinate,
                            "target": target,
                            "display": cell.value,
                            "tooltip": h.tooltip,
                            "internal": internal,
                        }
                    )
        return links

    def read_images(self, workbook: Workbook, sheet: str) -> list[JSONDict]:
        ws = workbook[sheet]
        images: list[JSONDict] = []
        for img in getattr(ws, "_images", []):
            anchor = getattr(img, "anchor", None)
            anchor_type = None
            cell = None
            offset = None
            if isinstance(anchor, str):
                anchor_type = "oneCell"
                cell = anchor
            from_anchor = getattr(anchor, "_from", None)
            if from_anchor is not None:
                anchor_type = "oneCell"
                col = getattr(from_anchor, "col", None)
                row = getattr(from_anchor, "row", None)
                if isinstance(col, int) and isinstance(row, int):
                    cell = f"{_col_letter(col + 1)}{row + 1}"
                col_off = getattr(from_anchor, "colOff", None)
                row_off = getattr(from_anchor, "rowOff", None)
                if col_off is not None and row_off is not None:
                    offset = [col_off, row_off]

            if getattr(anchor, "_to", None) is not None:
                anchor_type = "twoCell"
            images.append(
                {
                    "cell": cell,
                    "path": getattr(img, "path", None) or getattr(img, "_path", None),
                    "anchor": anchor_type,
                    "offset": offset,
                    "alt_text": getattr(img, "title", None),
                }
            )
        return images

    def read_pivot_tables(self, workbook: Workbook, sheet: str) -> list[JSONDict]:
        ws = workbook[sheet]
        pivots: list[JSONDict] = []
        pivot_list = getattr(ws, "_pivots", []) or []
        for pivot in pivot_list:
            source_range = None
            cache = getattr(pivot, "cache", None)
            cache_source = getattr(cache, "cacheSource", None) if cache is not None else None
            if cache_source is not None:
                worksheet_source = getattr(cache_source, "worksheetSource", None)
                ref = (
                    getattr(worksheet_source, "ref", None) if worksheet_source is not None else None
                )
                source_sheet = (
                    getattr(worksheet_source, "sheet", None)
                    if worksheet_source is not None
                    else None
                )
                if source_sheet and ref:
                    source_range = f"{source_sheet}!{ref}"
                elif ref:
                    source_range = ref
                else:
                    fallback = getattr(cache_source, "ref", None)
                    if fallback:
                        source_range = fallback
                    else:
                        source_range = str(cache_source)

            location = getattr(pivot, "location", None)
            target_cell = None
            if isinstance(location, str):
                target_cell = location
            elif location is not None:
                target_cell = getattr(location, "ref", None) or str(location)

            if target_cell and "!" not in target_cell:
                target_cell = f"{sheet}!{target_cell}"

            pivots.append(
                {
                    "name": getattr(pivot, "name", None),
                    "source_range": source_range,
                    "target_cell": target_cell,
                }
            )
        return pivots

    def read_comments(self, workbook: Workbook, sheet: str) -> list[JSONDict]:
        ws = workbook[sheet]
        comments: list[JSONDict] = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    comments.append(
                        {
                            "cell": cell.coordinate,
                            "text": cell.comment.text,
                            "author": cell.comment.author,
                            "threaded": False,
                        }
                    )
        return comments

    def read_freeze_panes(self, workbook: Workbook, sheet: str) -> JSONDict:
        ws = workbook[sheet]
        result: JSONDict = {}
        if ws.freeze_panes:
            freeze = ws.freeze_panes
            result["mode"] = "freeze"
            coord = getattr(freeze, "coordinate", None)
            result["top_left_cell"] = coord if coord else str(freeze)
        pane = getattr(ws.sheet_view, "pane", None)
        if pane and pane.state == "split" and (pane.xSplit or pane.ySplit):
            result["mode"] = "split"
            x_val = int(pane.xSplit) if pane.xSplit is not None else None
            y_val = int(pane.ySplit) if pane.ySplit is not None else None
            # xlsxwriter stores split values as twips (assuming default Calibri 11pt):
            #   y_twips = 20 * rows + 300,  x_twips = 180 * cols + 390
            # Convert back to logical row/col counts when values exceed a threshold
            # (logical values are small integers; twip values start at >=300).
            twips_x_offset = 390
            twips_x_factor = 180
            twips_y_offset = 300
            twips_y_factor = 20
            twips_conversion_threshold = 100
            if x_val is not None and x_val > twips_conversion_threshold:
                x_val = round((x_val - twips_x_offset) / twips_x_factor)
            if y_val is not None and y_val > twips_conversion_threshold:
                y_val = round((y_val - twips_y_offset) / twips_y_factor)
            result["x_split"] = x_val
            result["y_split"] = y_val
            if pane.topLeftCell:
                result["top_left_cell"] = pane.topLeftCell
            if pane.activePane:
                result["active_pane"] = pane.activePane
        return result

    # =========================================================================
    # Write Operations
    # =========================================================================

    def create_workbook(self) -> Workbook:
        """Create a new workbook."""
        wb = Workbook()
        # Remove default sheet to allow explicit sheet creation
        if wb.sheetnames:
            default_sheet = wb.active
            if default_sheet is not None:
                wb.remove(default_sheet)
        return wb

    def add_sheet(self, workbook: Workbook, name: str) -> None:
        """Add a new sheet to a workbook."""
        workbook.create_sheet(name)

    def write_cell_value(
        self,
        workbook: Workbook,
        sheet: str,
        cell: str,
        value: CellValue,
    ) -> None:
        """Write a value to a cell."""
        ws = workbook[sheet]
        c: Cell = ws[cell]

        if value.type == CellType.BLANK:
            c.value = None
        elif value.type == CellType.FORMULA:
            c.value = value.formula or value.value
        elif value.type == CellType.ERROR:
            # Write a formula that produces the error
            error_formulas = {
                "#DIV/0!": "=1/0",
                "#N/A": "=NA()",
                "#VALUE!": '="text"+1',
                "#REF!": "=#REF!",
                "#NAME?": "=_undefined_name_",
                "#NUM!": "=SQRT(-1)",
                "#NULL!": "=A1:A2 B1:B2",
            }
            c.value = error_formulas.get(value.value, value.value)
        else:
            c.value = value.value

    def write_cell_format(
        self,
        workbook: Workbook,
        sheet: str,
        cell: str,
        format: CellFormat,
    ) -> None:
        """Apply formatting to a cell."""
        ws = workbook[sheet]
        c: Cell = ws[cell]

        # Build font kwargs
        font_kwargs: dict[str, Any] = {}

        if format.bold is not None:
            font_kwargs["bold"] = format.bold
        if format.italic is not None:
            font_kwargs["italic"] = format.italic
        if format.underline is not None:
            font_kwargs["underline"] = format.underline
        if format.strikethrough is not None:
            font_kwargs["strike"] = format.strikethrough
        if format.font_name is not None:
            font_kwargs["name"] = format.font_name
        if format.font_size is not None:
            font_kwargs["size"] = format.font_size
        if format.font_color is not None:
            from openpyxl.styles import Color

            # Remove # prefix if present
            hex_color = format.font_color.lstrip("#")
            font_kwargs["color"] = Color(rgb=f"FF{hex_color}")

        if font_kwargs:
            c.font = Font(**font_kwargs)

        # Apply background color
        if format.bg_color is not None:
            hex_color = format.bg_color.lstrip("#")
            c.fill = PatternFill(
                start_color=f"FF{hex_color}",
                end_color=f"FF{hex_color}",
                fill_type="solid",
            )

        if format.number_format is not None:
            c.number_format = format.number_format

        align_kwargs: dict[str, Any] = {}
        if format.h_align is not None:
            align_kwargs["horizontal"] = format.h_align
        if format.v_align is not None:
            align_kwargs["vertical"] = format.v_align
        if format.wrap is not None:
            align_kwargs["wrap_text"] = format.wrap
        if format.rotation is not None:
            align_kwargs["text_rotation"] = format.rotation
        if format.indent is not None:
            align_kwargs["indent"] = format.indent
        if align_kwargs:
            c.alignment = Alignment(**align_kwargs)

    def write_cell_border(
        self,
        workbook: Workbook,
        sheet: str,
        cell: str,
        border: BorderInfo,
    ) -> None:
        """Apply border to a cell."""
        ws = workbook[sheet]
        c: Cell = ws[cell]

        def make_side(edge: BorderEdge | None) -> Side:
            if edge is None:
                return Side()

            # Map our style to openpyxl style
            style_map = {
                BorderStyle.NONE: None,
                BorderStyle.THIN: "thin",
                BorderStyle.MEDIUM: "medium",
                BorderStyle.THICK: "thick",
                BorderStyle.DOUBLE: "double",
                BorderStyle.DASHED: "dashed",
                BorderStyle.DOTTED: "dotted",
                BorderStyle.HAIR: "hair",
                BorderStyle.MEDIUM_DASHED: "mediumDashed",
                BorderStyle.DASH_DOT: "dashDot",
                BorderStyle.MEDIUM_DASH_DOT: "mediumDashDot",
                BorderStyle.DASH_DOT_DOT: "dashDotDot",
                BorderStyle.MEDIUM_DASH_DOT_DOT: "mediumDashDotDot",
                BorderStyle.SLANT_DASH_DOT: "slantDashDot",
            }

            style = style_map.get(edge.style)
            if style is None:
                return Side()

            hex_color = edge.color.lstrip("#")
            from openpyxl.styles import Color

            return Side(style=style, color=Color(rgb=f"FF{hex_color}"))

        # Determine diagonal settings
        diagonal_side = Side()
        diagonal_up = False
        diagonal_down = False

        if border.diagonal_up:
            diagonal_side = make_side(border.diagonal_up)
            diagonal_up = True
        if border.diagonal_down:
            diagonal_side = make_side(border.diagonal_down)
            diagonal_down = True

        c.border = Border(
            left=make_side(border.left),
            right=make_side(border.right),
            top=make_side(border.top),
            bottom=make_side(border.bottom),
            diagonal=diagonal_side,
            diagonalUp=diagonal_up,
            diagonalDown=diagonal_down,
        )

    def save_workbook(self, workbook: Workbook, path: Path) -> None:
        """Save a workbook to a file."""
        workbook.save(str(path))

    def set_row_height(
        self,
        workbook: Workbook,
        sheet: str,
        row: int,
        height: float,
    ) -> None:
        ws = workbook[sheet]
        ws.row_dimensions[row].height = height

    def set_column_width(
        self,
        workbook: Workbook,
        sheet: str,
        column: str,
        width: float,
    ) -> None:
        ws = workbook[sheet]
        ws.column_dimensions[column].width = width

    # =========================================================================
    # Tier 2 Write Operations
    # =========================================================================

    def merge_cells(self, workbook: Workbook, sheet: str, cell_range: str) -> None:
        ws = workbook[sheet]
        ws.merge_cells(cell_range)

    def add_conditional_format(self, workbook: Workbook, sheet: str, rule: JSONDict) -> None:
        ws = workbook[sheet]
        cf = rule.get("cf_rule", rule)
        range_ref = cf.get("range")
        rule_type = cf.get("rule_type")
        formula = cf.get("formula")
        operator = cf.get("operator")
        stop_if_true = cf.get("stop_if_true", False)
        fmt = cf.get("format") or {}

        fill = None
        font = None
        if fmt.get("bg_color"):
            hex_color = fmt["bg_color"].lstrip("#")
            fill = PatternFill(
                start_color=f"FF{hex_color}", end_color=f"FF{hex_color}", fill_type="solid"
            )
        if fmt.get("font_color"):
            hex_color = fmt["font_color"].lstrip("#")
            font = Font(color=f"FF{hex_color}")

        rule_obj = None
        if rule_type in ("cellIs", "cellIsRule"):
            from openpyxl.formatting.rule import CellIsRule

            rule_obj = CellIsRule(
                operator=operator,
                formula=[formula],
                fill=fill,
                font=font,
                stopIfTrue=stop_if_true,
            )
        elif rule_type in ("expression", "formula"):
            rule_obj = FormulaRule(
                formula=[formula],
                fill=fill,
                font=font,
                stopIfTrue=stop_if_true,
            )
        elif rule_type == "colorScale":
            rule_obj = ColorScaleRule(
                start_type="min",
                start_color="FFAA0000",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFFFFF00",
                end_type="max",
                end_color="FF00AA00",
            )
        elif rule_type == "dataBar":
            rule_obj = DataBarRule(
                start_type="min", end_type="max", color="FF638EC6", showValue=True
            )

        if rule_obj is not None:
            priority = cf.get("priority")
            if priority is not None:
                rule_obj.priority = priority
            ws.conditional_formatting.add(range_ref, rule_obj)

    def add_data_validation(self, workbook: Workbook, sheet: str, validation: JSONDict) -> None:
        ws = workbook[sheet]
        v = validation.get("validation", validation)
        dv = DataValidation(
            type=v.get("validation_type"),
            operator=v.get("operator"),
            formula1=v.get("formula1"),
            formula2=v.get("formula2"),
            allow_blank=v.get("allow_blank"),
            showInputMessage=v.get("show_input"),
            showErrorMessage=v.get("show_error"),
            promptTitle=v.get("prompt_title"),
            prompt=v.get("prompt"),
            errorTitle=v.get("error_title"),
            error=v.get("error"),
        )
        ws.add_data_validation(dv)
        dv.add(v.get("range"))

    def add_hyperlink(self, workbook: Workbook, sheet: str, link: JSONDict) -> None:
        ws = workbook[sheet]
        data = link.get("hyperlink", link)
        cell = data.get("cell")
        target = data.get("target")
        display = data.get("display")
        tooltip = data.get("tooltip")
        internal = data.get("internal")
        c_obj = ws[cell]
        if isinstance(c_obj, tuple):
            first = c_obj[0]
            if isinstance(first, tuple):
                c = first[0]
            else:
                c = first
        else:
            c = c_obj
        if display is not None:
            c.value = display
        if internal:
            location = target.lstrip("#") if target else target
            c.hyperlink = Hyperlink(ref=cell, location=location)
        else:
            c.hyperlink = target
        if c.hyperlink and tooltip is not None:
            c.hyperlink.tooltip = tooltip

    def add_image(self, workbook: Workbook, sheet: str, image: JSONDict) -> None:
        ws = workbook[sheet]
        data = image.get("image", image)
        path = data.get("path")
        cell = data.get("cell")
        if not path or not cell:
            return
        img = Image(path)
        ws.add_image(img, cell)

    def add_pivot_table(self, workbook: Workbook, sheet: str, pivot: JSONDict) -> None:
        raise NotImplementedError("openpyxl does not support pivot table creation")

    def add_comment(self, workbook: Workbook, sheet: str, comment: JSONDict) -> None:
        ws = workbook[sheet]
        data = comment.get("comment", comment)
        cell = data.get("cell")
        text = data.get("text")
        author = data.get("author") or ""
        if cell and text is not None:
            c_obj = ws[cell]
            if isinstance(c_obj, tuple):
                first = c_obj[0]
                if isinstance(first, tuple):
                    c = first[0]
                else:
                    c = first
            else:
                c = c_obj
            c.comment = Comment(text, author)

    def set_freeze_panes(self, workbook: Workbook, sheet: str, settings: JSONDict) -> None:
        ws = workbook[sheet]
        cfg = settings.get("freeze", settings)
        mode = cfg.get("mode")
        if mode == "freeze":
            ws.freeze_panes = cfg.get("top_left_cell")
        elif mode == "split":
            from openpyxl.worksheet.views import Pane

            ws.freeze_panes = None
            pane = ws.sheet_view.pane
            if pane is None:
                pane = Pane()
                ws.sheet_view.pane = pane
            if cfg.get("x_split") is not None:
                pane.xSplit = cfg["x_split"]
            if cfg.get("y_split") is not None:
                pane.ySplit = cfg["y_split"]
            if cfg.get("top_left_cell") is not None:
                pane.topLeftCell = cfg["top_left_cell"]
            if cfg.get("active_pane") is not None:
                pane.activePane = cfg["active_pane"]
            pane.state = "split"
