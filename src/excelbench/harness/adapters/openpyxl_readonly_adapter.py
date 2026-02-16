"""Adapter for openpyxl in read-only mode.

Uses ``openpyxl.load_workbook(read_only=True)`` which streams data via
lazy-loading iterators instead of building the full DOM.  Some metadata
(merged cells, conditional formatting, data validations) may not be
available — measuring the read-mode fidelity difference vs. full mode.
"""

from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

from excelbench.harness.adapters.base import ReadOnlyAdapter
from excelbench.models import (
    BorderInfo,
    CellFormat,
    CellType,
    CellValue,
    LibraryInfo,
)

JSONDict = dict[str, Any]


def _get_version() -> str:
    return str(openpyxl.__version__)


class OpenpyxlReadonlyAdapter(ReadOnlyAdapter):
    """openpyxl with ``read_only=True`` (streaming reads).

    Read-only mode returns ``ReadOnlyCell`` objects which lack some
    formatting and structural metadata.  Write operations are not
    supported and raise ``NotImplementedError`` via :class:`ReadOnlyAdapter`.
    """

    @property
    def info(self) -> LibraryInfo:
        return LibraryInfo(
            name="openpyxl-readonly",
            version=_get_version(),
            language="python",
            capabilities={"read"},
        )

    @property
    def supported_read_extensions(self) -> set[str]:
        return {".xlsx"}

    # =========================================================================
    # Read Operations
    # =========================================================================

    def open_workbook(self, path: Path) -> Any:
        return openpyxl.load_workbook(str(path), data_only=False, read_only=True)

    def close_workbook(self, workbook: Any) -> None:
        workbook.close()

    def get_sheet_names(self, workbook: Any) -> list[str]:
        return [str(name) for name in workbook.sheetnames]

    def read_sheet_values(
        self,
        workbook: Any,
        sheet: str,
        cell_range: str | None = None,
    ) -> list[list[CellValue]]:
        """Bulk read a range of values in streaming mode.

        Optional helper used by performance workloads.
        """
        ws = workbook[sheet]

        if cell_range:
            import re

            clean = cell_range.replace("$", "").upper()
            if ":" in clean:
                a, b = clean.split(":", 1)
            else:
                a, b = clean, clean

            def _cell_to_rc(cell: str) -> tuple[int, int]:
                m = re.match(r"([A-Z]+)(\d+)", cell)
                if not m:
                    return 1, 1
                col_str, row_str = m.groups()
                row = int(row_str)
                col = 0
                for ch in col_str:
                    col = col * 26 + (ord(ch) - ord("A") + 1)
                return row, col

            r0, c0 = _cell_to_rc(a)
            r1, c1 = _cell_to_rc(b)
            if r1 < r0:
                r0, r1 = r1, r0
            if c1 < c0:
                c0, c1 = c1, c0

            rows = ws.iter_rows(min_row=r0, max_row=r1, min_col=c0, max_col=c1)
        else:
            rows = ws.iter_rows()

        out: list[list[CellValue]] = []
        for row in rows:
            out.append([self._classify_value(c) for c in row])
        return out

    def read_sheet_values_raw(
        self,
        workbook: Any,
        sheet: str,
        cell_range: str | None = None,
    ) -> list[list[Any]]:
        """Return raw ReadOnlyCell rows without CellValue conversion."""
        ws = workbook[sheet]
        if cell_range:
            import re

            clean = cell_range.replace("$", "").upper()
            if ":" in clean:
                a, b = clean.split(":", 1)
            else:
                a, b = clean, clean

            def _cell_to_rc(ref: str) -> tuple[int, int]:
                m = re.match(r"([A-Z]+)(\d+)", ref)
                if not m:
                    return 1, 1
                col_str, row_str = m.groups()
                row = int(row_str)
                col = 0
                for ch in col_str:
                    col = col * 26 + (ord(ch) - ord("A") + 1)
                return row, col

            r0, c0 = _cell_to_rc(a)
            r1, c1 = _cell_to_rc(b)
            if r1 < r0:
                r0, r1 = r1, r0
            if c1 < c0:
                c0, c1 = c1, c0
            return [list(row) for row in ws.iter_rows(
                min_row=r0, max_row=r1, min_col=c0, max_col=c1
            )]
        return [list(row) for row in ws.iter_rows()]

    def read_cell_value(
        self,
        workbook: Any,
        sheet: str,
        cell: str,
    ) -> CellValue:
        ws = workbook[sheet]

        # ReadOnlyWorksheet doesn't support ws[cell] random access directly.
        # We need to use iter_rows to reach the target cell.
        import re

        match = re.match(r"([A-Z]+)(\d+)", cell.upper())
        if not match:
            return CellValue(type=CellType.BLANK)

        col_str, row_str = match.groups()
        target_row = int(row_str)
        target_col = 0
        for char in col_str:
            target_col = target_col * 26 + (ord(char) - ord("A") + 1)

        # iter_rows with specific range for efficiency
        for row in ws.iter_rows(
            min_row=target_row,
            max_row=target_row,
            min_col=target_col,
            max_col=target_col,
        ):
            if row:
                c = row[0]
                return self._classify_value(c)

        return CellValue(type=CellType.BLANK)

    @staticmethod
    def _classify_value(c: Any) -> CellValue:
        value = c.value

        if value is None:
            return CellValue(type=CellType.BLANK)

        if isinstance(value, bool):
            return CellValue(type=CellType.BOOLEAN, value=value)

        if isinstance(value, (int, float)):
            return CellValue(type=CellType.NUMBER, value=value)

        if isinstance(value, date) and not isinstance(value, datetime):
            return CellValue(type=CellType.DATE, value=value)

        if isinstance(value, datetime):
            if (
                value.hour == 0
                and value.minute == 0
                and value.second == 0
                and value.microsecond == 0
            ):
                return CellValue(type=CellType.DATE, value=value.date())
            return CellValue(type=CellType.DATETIME, value=value)

        if isinstance(value, str):
            if value in ("#N/A", "#NULL!", "#NAME?", "#REF!"):
                return CellValue(type=CellType.ERROR, value=value)
            if value.startswith("#") and value.endswith("!"):
                return CellValue(type=CellType.ERROR, value=value)

            # Check formula via data_type attribute
            data_type = getattr(c, "data_type", None)
            if data_type == "f" or value.startswith("="):
                formula_str = value if value.startswith("=") else f"={value}"
                error_formula_map = {
                    "=1/0": "#DIV/0!",
                    "=NA()": "#N/A",
                    '="text"+1': "#VALUE!",
                }
                if formula_str in error_formula_map:
                    return CellValue(type=CellType.ERROR, value=error_formula_map[formula_str])
                return CellValue(type=CellType.FORMULA, value=value, formula=formula_str)

            return CellValue(type=CellType.STRING, value=value)

        return CellValue(type=CellType.STRING, value=str(value))

    def read_cell_format(self, workbook: Any, sheet: str, cell: str) -> CellFormat:
        # ReadOnlyCell has limited formatting support — return defaults
        return CellFormat()

    def read_cell_border(self, workbook: Any, sheet: str, cell: str) -> BorderInfo:
        return BorderInfo()

    def read_row_height(self, workbook: Any, sheet: str, row: int) -> float | None:
        return None

    def read_column_width(self, workbook: Any, sheet: str, column: str) -> float | None:
        return None

    # =========================================================================
    # Tier 2 Read Operations
    # =========================================================================

    def read_merged_ranges(self, workbook: Any, sheet: str) -> list[str]:
        # merged_cells is available in read-only mode
        ws = workbook[sheet]
        try:
            return [str(rng) for rng in ws.merged_cells.ranges]
        except (AttributeError, TypeError):
            return []

    def read_conditional_formats(self, workbook: Any, sheet: str) -> list[JSONDict]:
        return []

    def read_data_validations(self, workbook: Any, sheet: str) -> list[JSONDict]:
        return []

    def read_hyperlinks(self, workbook: Any, sheet: str) -> list[JSONDict]:
        return []

    def read_images(self, workbook: Any, sheet: str) -> list[JSONDict]:
        return []

    def read_pivot_tables(self, workbook: Any, sheet: str) -> list[JSONDict]:
        return []

    def read_comments(self, workbook: Any, sheet: str) -> list[JSONDict]:
        return []

    def read_freeze_panes(self, workbook: Any, sheet: str) -> JSONDict:
        return {}
