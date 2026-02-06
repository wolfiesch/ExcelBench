"""Generator for border test cases."""

import sys
from pathlib import Path
from typing import TypedDict

import xlwings as xw

from excelbench.generator.base import FeatureGenerator
from excelbench.models import TestCase


class _BorderEdgeSpec(TypedDict, total=False):
    line_style: int
    weight: int
    color: tuple[int, int, int]


BorderEdges = dict[str, _BorderEdgeSpec]


class _BorderOp(TypedDict):
    row: int
    edges: BorderEdges


# Excel border style constants (from xlwings/Excel API)
class XlBorderWeight:
    HAIRLINE = 1
    THIN = 2
    MEDIUM = -4138
    THICK = 4


class XlLineStyle:
    CONTINUOUS = 1
    DASH = -4115
    DASH_DOT = 4
    DASH_DOT_DOT = 5
    DOT = -4118
    DOUBLE = -4119
    NONE = -4142
    SLANT_DASH_DOT = 13


class XlBordersIndex:
    EDGE_LEFT = 7
    EDGE_TOP = 8
    EDGE_BOTTOM = 9
    EDGE_RIGHT = 10
    INSIDE_VERTICAL = 11
    INSIDE_HORIZONTAL = 12
    DIAGONAL_DOWN = 5
    DIAGONAL_UP = 6


class BordersGenerator(FeatureGenerator):
    """Generates test cases for cell borders.

    Tests: styles, weights, colors, positions, diagonals.
    """

    feature_name = "borders"
    tier = 1
    filename = "07_borders.xlsx"

    def __init__(self) -> None:
        self._use_openpyxl = sys.platform == "darwin"
        self._border_ops: list[_BorderOp] = []

    def generate(self, sheet: xw.Sheet) -> list[TestCase]:
        """Generate border test cases."""
        self.setup_header(sheet)

        test_cases = []
        row = 2

        # Border styles
        test_cases.append(self._test_thin_all(sheet, row))
        row += 1

        test_cases.append(self._test_medium_all(sheet, row))
        row += 1

        test_cases.append(self._test_thick_all(sheet, row))
        row += 1

        test_cases.append(self._test_double(sheet, row))
        row += 1

        test_cases.append(self._test_dashed(sheet, row))
        row += 1

        test_cases.append(self._test_dotted(sheet, row))
        row += 1

        test_cases.append(self._test_dash_dot(sheet, row))
        row += 1

        test_cases.append(self._test_dash_dot_dot(sheet, row))
        row += 1

        # Individual edges
        test_cases.append(self._test_top_only(sheet, row))
        row += 1

        test_cases.append(self._test_bottom_only(sheet, row))
        row += 1

        test_cases.append(self._test_left_only(sheet, row))
        row += 1

        test_cases.append(self._test_right_only(sheet, row))
        row += 1

        # Diagonal borders
        test_cases.append(self._test_diagonal_up(sheet, row))
        row += 1

        test_cases.append(self._test_diagonal_down(sheet, row))
        row += 1

        test_cases.append(self._test_diagonal_both(sheet, row))
        row += 1

        # Colors
        test_cases.append(self._test_color_red(sheet, row))
        row += 1

        test_cases.append(self._test_color_blue(sheet, row))
        row += 1

        test_cases.append(self._test_color_custom(sheet, row))
        row += 1

        # Mixed edges
        test_cases.append(self._test_mixed_styles(sheet, row))
        row += 1

        test_cases.append(self._test_mixed_colors(sheet, row))
        row += 1

        return test_cases

    def _set_all_borders(
        self,
        cell: xw.Range,
        weight: int = XlBorderWeight.THIN,
        line_style: int = XlLineStyle.CONTINUOUS,
        color: tuple[int, int, int] | None = None,
    ) -> None:
        """Set all four edges of a cell border."""

        def _spec() -> _BorderEdgeSpec:
            spec: _BorderEdgeSpec = {"line_style": line_style, "weight": weight}
            if color is not None:
                spec["color"] = color
            return spec

        edges: BorderEdges = {
            "top": _spec(),
            "bottom": _spec(),
            "left": _spec(),
            "right": _spec(),
        }
        self._apply_border_edges(cell, edges)

    def _rgb_to_int(self, rgb: tuple[int, int, int]) -> int:
        """Convert RGB tuple to Excel color integer."""
        return rgb[0] + (rgb[1] * 256) + (rgb[2] * 256 * 256)

    def _test_thin_all(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Border - thin all edges"
        expected = {"border_style": "thin", "border_color": "#000000"}

        self.write_test_case(sheet, row, label, expected)
        cell = sheet.range(f"B{row}")
        cell.value = "Thin"
        self._set_all_borders(cell, XlBorderWeight.THIN)

        return TestCase(id="thin_all", label=label, row=row, expected=expected)

    def _test_medium_all(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Border - medium all edges"
        expected = {"border_style": "medium", "border_color": "#000000"}

        self.write_test_case(sheet, row, label, expected)
        cell = sheet.range(f"B{row}")
        cell.value = "Medium"
        self._set_all_borders(cell, XlBorderWeight.MEDIUM)

        return TestCase(id="medium_all", label=label, row=row, expected=expected)

    def _test_thick_all(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Border - thick all edges"
        expected = {"border_style": "thick", "border_color": "#000000"}

        self.write_test_case(sheet, row, label, expected)
        cell = sheet.range(f"B{row}")
        cell.value = "Thick"
        self._set_all_borders(cell, XlBorderWeight.THICK)

        return TestCase(id="thick_all", label=label, row=row, expected=expected)

    def _test_double(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Border - double line"
        expected = {"border_style": "double", "border_color": "#000000"}

        self.write_test_case(sheet, row, label, expected)
        cell = sheet.range(f"B{row}")
        cell.value = "Double"
        self._set_all_borders(cell, line_style=XlLineStyle.DOUBLE)

        return TestCase(id="double", label=label, row=row, expected=expected)

    def _test_dashed(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Border - dashed"
        expected = {"border_style": "dashed", "border_color": "#000000"}

        self.write_test_case(sheet, row, label, expected)
        cell = sheet.range(f"B{row}")
        cell.value = "Dashed"
        self._set_all_borders(cell, line_style=XlLineStyle.DASH)

        return TestCase(id="dashed", label=label, row=row, expected=expected)

    def _test_dotted(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Border - dotted"
        expected = {"border_style": "dotted", "border_color": "#000000"}

        self.write_test_case(sheet, row, label, expected)
        cell = sheet.range(f"B{row}")
        cell.value = "Dotted"
        self._set_all_borders(cell, line_style=XlLineStyle.DOT)

        return TestCase(id="dotted", label=label, row=row, expected=expected)

    def _test_dash_dot(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Border - dash-dot"
        expected = {"border_style": "dashDot", "border_color": "#000000"}

        self.write_test_case(sheet, row, label, expected)
        cell = sheet.range(f"B{row}")
        cell.value = "Dash-Dot"
        self._set_all_borders(cell, line_style=XlLineStyle.DASH_DOT)

        return TestCase(id="dash_dot", label=label, row=row, expected=expected)

    def _test_dash_dot_dot(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Border - dash-dot-dot"
        expected = {"border_style": "dashDotDot", "border_color": "#000000"}

        self.write_test_case(sheet, row, label, expected)
        cell = sheet.range(f"B{row}")
        cell.value = "Dash-Dot-Dot"
        self._set_all_borders(cell, line_style=XlLineStyle.DASH_DOT_DOT)

        return TestCase(id="dash_dot_dot", label=label, row=row, expected=expected)

    def _test_top_only(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Border - top only"
        expected = {
            "border_top": "thin",
            "border_bottom": None,
            "border_left": None,
            "border_right": None,
        }

        self.write_test_case(sheet, row, label, expected)
        cell = sheet.range(f"B{row}")
        cell.value = "Top Only"
        self._apply_border_edges(
            cell,
            {"top": {"line_style": XlLineStyle.CONTINUOUS, "weight": XlBorderWeight.THIN}},
        )

        return TestCase(id="top_only", label=label, row=row, expected=expected)

    def _test_bottom_only(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Border - bottom only"
        expected = {
            "border_top": None,
            "border_bottom": "thin",
            "border_left": None,
            "border_right": None,
        }

        self.write_test_case(sheet, row, label, expected)
        cell = sheet.range(f"B{row}")
        cell.value = "Bottom Only"
        self._apply_border_edges(
            cell,
            {"bottom": {"line_style": XlLineStyle.CONTINUOUS, "weight": XlBorderWeight.THIN}},
        )

        return TestCase(id="bottom_only", label=label, row=row, expected=expected)

    def _test_left_only(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Border - left only"
        expected = {
            "border_top": None,
            "border_bottom": None,
            "border_left": "thin",
            "border_right": None,
        }

        self.write_test_case(sheet, row, label, expected)
        cell = sheet.range(f"B{row}")
        cell.value = "Left Only"
        self._apply_border_edges(
            cell,
            {"left": {"line_style": XlLineStyle.CONTINUOUS, "weight": XlBorderWeight.THIN}},
        )

        return TestCase(id="left_only", label=label, row=row, expected=expected)

    def _test_right_only(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Border - right only"
        expected = {
            "border_top": None,
            "border_bottom": None,
            "border_left": None,
            "border_right": "thin",
        }

        self.write_test_case(sheet, row, label, expected)
        cell = sheet.range(f"B{row}")
        cell.value = "Right Only"
        self._apply_border_edges(
            cell,
            {"right": {"line_style": XlLineStyle.CONTINUOUS, "weight": XlBorderWeight.THIN}},
        )

        return TestCase(id="right_only", label=label, row=row, expected=expected)

    def _test_diagonal_up(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Border - diagonal up"
        expected = {"border_diagonal_up": "thin"}

        self.write_test_case(sheet, row, label, expected)
        cell = sheet.range(f"B{row}")
        cell.value = "Diag Up"
        self._apply_border_edges(
            cell,
            {
                "diagonal_up": {
                    "line_style": XlLineStyle.CONTINUOUS,
                    "weight": XlBorderWeight.THIN,
                }
            },
        )

        return TestCase(id="diagonal_up", label=label, row=row, expected=expected)

    def _test_diagonal_down(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Border - diagonal down"
        expected = {"border_diagonal_down": "thin"}

        self.write_test_case(sheet, row, label, expected)
        cell = sheet.range(f"B{row}")
        cell.value = "Diag Down"
        self._apply_border_edges(
            cell,
            {
                "diagonal_down": {
                    "line_style": XlLineStyle.CONTINUOUS,
                    "weight": XlBorderWeight.THIN,
                }
            },
        )

        return TestCase(id="diagonal_down", label=label, row=row, expected=expected)

    def _test_diagonal_both(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Border - diagonal both"
        expected = {"border_diagonal_up": "thin", "border_diagonal_down": "thin"}

        self.write_test_case(sheet, row, label, expected)
        cell = sheet.range(f"B{row}")
        cell.value = "Both Diag"
        self._apply_border_edges(
            cell,
            {
                "diagonal_up": {
                    "line_style": XlLineStyle.CONTINUOUS,
                    "weight": XlBorderWeight.THIN,
                },
                "diagonal_down": {
                    "line_style": XlLineStyle.CONTINUOUS,
                    "weight": XlBorderWeight.THIN,
                },
            },
        )

        return TestCase(id="diagonal_both", label=label, row=row, expected=expected)

    def _test_color_red(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Border - red color"
        expected = {"border_style": "thin", "border_color": "#FF0000"}

        self.write_test_case(sheet, row, label, expected)
        cell = sheet.range(f"B{row}")
        cell.value = "Red Border"
        self._set_all_borders(cell, XlBorderWeight.THIN, color=(255, 0, 0))

        return TestCase(id="color_red", label=label, row=row, expected=expected)

    def _test_color_blue(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Border - blue color"
        expected = {"border_style": "thin", "border_color": "#0000FF"}

        self.write_test_case(sheet, row, label, expected)
        cell = sheet.range(f"B{row}")
        cell.value = "Blue Border"
        self._set_all_borders(cell, XlBorderWeight.THIN, color=(0, 0, 255))

        return TestCase(id="color_blue", label=label, row=row, expected=expected)

    def _test_color_custom(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Border - custom color (#8B4513)"
        expected = {"border_style": "thin", "border_color": "#8B4513"}

        self.write_test_case(sheet, row, label, expected)
        cell = sheet.range(f"B{row}")
        cell.value = "Custom Color"
        self._set_all_borders(cell, XlBorderWeight.THIN, color=(139, 69, 19))

        return TestCase(id="color_custom", label=label, row=row, expected=expected)

    def _test_mixed_styles(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Border - mixed styles per edge"
        expected = {
            "border_top": "thick",
            "border_bottom": "thin",
            "border_left": "medium",
            "border_right": "dashed",
        }

        self.write_test_case(sheet, row, label, expected)
        cell = sheet.range(f"B{row}")
        cell.value = "Mixed Styles"
        self._apply_border_edges(
            cell,
            {
                "top": {"line_style": XlLineStyle.CONTINUOUS, "weight": XlBorderWeight.THICK},
                "bottom": {"line_style": XlLineStyle.CONTINUOUS, "weight": XlBorderWeight.THIN},
                "left": {"line_style": XlLineStyle.CONTINUOUS, "weight": XlBorderWeight.MEDIUM},
                "right": {"line_style": XlLineStyle.DASH, "weight": XlBorderWeight.THIN},
            },
        )

        return TestCase(id="mixed_styles", label=label, row=row, expected=expected)

    def _test_mixed_colors(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Border - mixed colors per edge"
        expected = {
            "border_top_color": "#FF0000",
            "border_bottom_color": "#00FF00",
            "border_left_color": "#0000FF",
            "border_right_color": "#FFFF00",
        }

        self.write_test_case(sheet, row, label, expected)
        cell = sheet.range(f"B{row}")
        cell.value = "Mixed Colors"
        self._apply_border_edges(
            cell,
            {
                "top": {
                    "line_style": XlLineStyle.CONTINUOUS,
                    "weight": XlBorderWeight.THIN,
                    "color": (255, 0, 0),
                },
                "bottom": {
                    "line_style": XlLineStyle.CONTINUOUS,
                    "weight": XlBorderWeight.THIN,
                    "color": (0, 255, 0),
                },
                "left": {
                    "line_style": XlLineStyle.CONTINUOUS,
                    "weight": XlBorderWeight.THIN,
                    "color": (0, 0, 255),
                },
                "right": {
                    "line_style": XlLineStyle.CONTINUOUS,
                    "weight": XlBorderWeight.THIN,
                    "color": (255, 255, 0),
                },
            },
        )

        return TestCase(id="mixed_colors", label=label, row=row, expected=expected)

    def _edge_index(self, name: str) -> int:
        mapping = {
            "top": XlBordersIndex.EDGE_TOP,
            "bottom": XlBordersIndex.EDGE_BOTTOM,
            "left": XlBordersIndex.EDGE_LEFT,
            "right": XlBordersIndex.EDGE_RIGHT,
            "diagonal_up": XlBordersIndex.DIAGONAL_UP,
            "diagonal_down": XlBordersIndex.DIAGONAL_DOWN,
        }
        return mapping[name]

    def _apply_border_edges(self, cell: xw.Range, edges: BorderEdges) -> None:
        if self._use_openpyxl:
            self._border_ops.append({"row": cell.row, "edges": edges})
            return
        for edge_name, spec in edges.items():
            border = cell.api.Borders(self._edge_index(edge_name))
            line_style = spec.get("line_style", XlLineStyle.CONTINUOUS)
            border.LineStyle = line_style
            weight = spec.get("weight")
            if weight is not None:
                border.Weight = weight
            color = spec.get("color")
            if color is not None:
                border.Color = self._rgb_to_int(color)

    def _openpyxl_style(self, line_style: int | None, weight: int | None) -> str | None:
        if line_style in (None, XlLineStyle.NONE):
            return None
        if line_style == XlLineStyle.CONTINUOUS:
            if weight == XlBorderWeight.THICK:
                return "thick"
            if weight == XlBorderWeight.MEDIUM:
                return "medium"
            if weight == XlBorderWeight.HAIRLINE:
                return "hair"
            return "thin"
        if line_style == XlLineStyle.DOUBLE:
            return "double"
        if line_style == XlLineStyle.DASH:
            return "dashed"
        if line_style == XlLineStyle.DOT:
            return "dotted"
        if line_style == XlLineStyle.DASH_DOT:
            return "dashDot"
        if line_style == XlLineStyle.DASH_DOT_DOT:
            return "dashDotDot"
        if line_style == XlLineStyle.SLANT_DASH_DOT:
            return "slantDashDot"
        return "thin"

    def _rgb_to_hex(self, rgb: tuple[int, int, int]) -> str:
        return f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"

    def post_process(self, output_path: Path) -> None:
        if not self._use_openpyxl or not self._border_ops:
            return
        from openpyxl import load_workbook
        from openpyxl.styles import Border, Side

        wb = load_workbook(output_path)
        ws = wb[self.feature_name]

        for op in self._border_ops:
            row = op["row"]
            edges = op["edges"]

            def side_for(name: str) -> Side | None:
                spec = edges.get(name)
                if not spec:
                    return None
                style = self._openpyxl_style(spec.get("line_style"), spec.get("weight"))
                if style is None:
                    return None
                color = spec.get("color")
                if isinstance(color, tuple):
                    return Side(style=style, color=self._rgb_to_hex(color))
                return Side(style=style)

            diag_up = edges.get("diagonal_up")
            diag_down = edges.get("diagonal_down")
            diag_spec = diag_up or diag_down
            diag_side = None
            if diag_spec:
                style = self._openpyxl_style(diag_spec.get("line_style"), diag_spec.get("weight"))
                if style is not None:
                    color = diag_spec.get("color")
                    if isinstance(color, tuple):
                        diag_side = Side(style=style, color=self._rgb_to_hex(color))
                    else:
                        diag_side = Side(style=style)

            border = Border(
                left=side_for("left"),
                right=side_for("right"),
                top=side_for("top"),
                bottom=side_for("bottom"),
                diagonal=diag_side,
                diagonalUp=diag_up is not None,
                diagonalDown=diag_down is not None,
            )
            ws[f"B{row}"].border = border

        wb.save(output_path)
