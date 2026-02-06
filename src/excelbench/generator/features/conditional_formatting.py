"""Generator for conditional formatting test cases (Tier 2)."""

import sys
from pathlib import Path

import xlwings as xw

from excelbench.generator.base import FeatureGenerator
from excelbench.models import ConditionalFormatSpec, Importance, TestCase


class ConditionalFormattingGenerator(FeatureGenerator):
    """Generates test cases for conditional formatting."""

    feature_name = "conditional_formatting"
    tier = 2
    filename = "11_conditional_formatting.xlsx"

    def __init__(self) -> None:
        self._use_openpyxl = sys.platform == "darwin"
        self._cf_ops: list[dict[str, object]] = []

    def generate(self, sheet: xw.Sheet) -> list[TestCase]:
        self.setup_header(sheet)

        # Seed values
        for i in range(2, 10):
            sheet.range(f"B{i}").value = i - 1

        # Reference sheet for formula-based CF
        ref_sheet = sheet.book.sheets.add("Ref")
        ref_sheet.range("A1").value = 10

        test_cases: list[TestCase] = []
        row = 2

        # 1) Cell value greater than
        label = "CF: cell > 5 (yellow fill)"
        target_range = "B2:B6"
        if not self._use_openpyxl:
            fc = sheet.range(target_range).api.FormatConditions.Add(
                Type=1,  # xlCellValue
                Operator=5,  # xlGreater
                Formula1="5",
            )
            fc.Interior.Color = 0x00FFFF  # Yellow in BGR
            fc.Priority = 1
        self._record_cf(
            kind="cellIs",
            range=target_range,
            operator="greaterThan",
            formula="5",
            priority=1,
            bg_color="#FFFF00",
        )
        expected = ConditionalFormatSpec(
            range=target_range,
            rule_type="cellIs",
            operator="greaterThan",
            formula="5",
            priority=1,
            format={"bg_color": "#FFFF00"},
        ).to_expected()
        self.write_test_case(sheet, row, label, expected)
        test_cases.append(TestCase(id="cf_cell_gt", label=label, row=row, expected=expected))
        row += 1

        # 2) Formula-based rule referencing another sheet
        label = "CF: formula rule with cross-sheet ref"
        target_range = "B2:B6"
        if not self._use_openpyxl:
            fc = sheet.range(target_range).api.FormatConditions.Add(
                Type=2,  # xlExpression
                Formula1="=Ref!$A$1>5",
            )
            fc.Interior.Color = 0xFF00FF  # Magenta in BGR
            fc.Priority = 2
        self._record_cf(
            kind="expression",
            range=target_range,
            formula="=Ref!$A$1>5",
            priority=2,
            bg_color="#FF00FF",
        )
        expected = ConditionalFormatSpec(
            range=target_range,
            rule_type="expression",
            formula="=Ref!$A$1>5",
            priority=2,
            format={"bg_color": "#FF00FF"},
        ).to_expected()
        self.write_test_case(sheet, row, label, expected)
        test_cases.append(
            TestCase(
                id="cf_formula_cross_sheet",
                label=label,
                row=row,
                expected=expected,
                importance=Importance.EDGE,
            )
        )
        row += 1

        # 2b) Text contains (formula-based)
        label = "CF: text contains"
        sheet.range("B2").value = "foo"
        target_range = "B2:B6"
        if not self._use_openpyxl:
            fc = sheet.range(target_range).api.FormatConditions.Add(
                Type=2,  # xlExpression
                Formula1='=ISNUMBER(SEARCH("foo",B2))',
            )
            fc.Interior.Color = 0x00FFFF  # Yellow in BGR
            fc.Priority = 3
        self._record_cf(
            kind="expression",
            range=target_range,
            formula='=ISNUMBER(SEARCH("foo",B2))',
            priority=3,
            bg_color="#FFFF00",
        )
        expected = ConditionalFormatSpec(
            range=target_range,
            rule_type="expression",
            formula='=ISNUMBER(SEARCH("foo",B2))',
            priority=3,
            format={"bg_color": "#FFFF00"},
        ).to_expected()
        self.write_test_case(sheet, row, label, expected)
        test_cases.append(
            TestCase(
                id="cf_text_contains",
                label=label,
                row=row,
                expected=expected,
                importance=Importance.EDGE,
            )
        )
        row += 1

        # 3) Data bar
        label = "CF: data bar"
        target_range = "B2:B6"
        if not self._use_openpyxl:
            fc = sheet.range(target_range).api.FormatConditions.AddDatabar()
            fc.MinPoint.Modify(1, 0)  # xlConditionValueNumber
            fc.MaxPoint.Modify(1, 10)
            fc.Priority = 4
        self._record_cf(
            kind="dataBar",
            range=target_range,
            min_value=0,
            max_value=10,
            priority=4,
        )
        expected = {
            "cf_rule": {
                "range": target_range,
                "rule_type": "dataBar",
                "priority": 4,
            }
        }
        self.write_test_case(sheet, row, label, expected)
        test_cases.append(
            TestCase(
                id="cf_data_bar",
                label=label,
                row=row,
                expected=expected,
                importance=Importance.EDGE,
            )
        )
        row += 1

        # 4) 3-color scale
        label = "CF: 3-color scale"
        target_range = "B2:B6"
        if not self._use_openpyxl:
            fc = sheet.range(target_range).api.FormatConditions.AddColorScale(ColorScaleType=3)
            fc.Priority = 5
        self._record_cf(kind="colorScale", range=target_range, priority=5)
        expected = {
            "cf_rule": {
                "range": target_range,
                "rule_type": "colorScale",
                "priority": 5,
            }
        }
        self.write_test_case(sheet, row, label, expected)
        test_cases.append(
            TestCase(
                id="cf_color_scale",
                label=label,
                row=row,
                expected=expected,
                importance=Importance.EDGE,
            )
        )
        row += 1

        # 5) Multiple rules same range + stop-if-true
        label = "CF: stop-if-true priority"
        target_range = "B7:B9"
        if not self._use_openpyxl:
            fc1 = sheet.range(target_range).api.FormatConditions.Add(
                Type=1,
                Operator=6,  # xlLess
                Formula1="3",
            )
            fc1.Interior.Color = 0x0000FF  # Red in BGR
            fc1.Priority = 1
            fc1.StopIfTrue = True
            fc2 = sheet.range(target_range).api.FormatConditions.Add(
                Type=1,
                Operator=7,  # xlGreaterEqual
                Formula1="7",
            )
            fc2.Interior.Color = 0x00FF00  # Green in BGR
            fc2.Priority = 2
        self._record_cf(
            kind="cellIs",
            range=target_range,
            operator="lessThan",
            formula="3",
            priority=1,
            bg_color="#FF0000",
            stop_if_true=True,
        )
        self._record_cf(
            kind="cellIs",
            range=target_range,
            operator="greaterThanOrEqual",
            formula="7",
            priority=2,
            bg_color="#00FF00",
        )

        expected = ConditionalFormatSpec(
            range=target_range,
            rule_type="cellIs",
            operator="lessThan",
            formula="3",
            priority=1,
            stop_if_true=True,
            format={"bg_color": "#FF0000"},
        ).to_expected()
        self.write_test_case(sheet, row, label, expected)
        test_cases.append(
            TestCase(
                id="cf_stop_if_true",
                label=label,
                row=row,
                expected=expected,
                importance=Importance.EDGE,
            )
        )

        return test_cases

    def _record_cf(self, **kwargs: object) -> None:
        self._cf_ops.append(kwargs)

    def _hex_no_hash(self, color: str) -> str:
        return color.lstrip("#")

    def post_process(self, output_path: Path) -> None:
        if not self._use_openpyxl or not self._cf_ops:
            return
        from openpyxl import load_workbook
        from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
        from openpyxl.styles import PatternFill

        wb = load_workbook(output_path)
        ws = wb[self.feature_name]

        for op in self._cf_ops:
            kind = op.get("kind")
            cf_range = op.get("range")
            if not isinstance(cf_range, str):
                continue
            bg_color = op.get("bg_color")
            fill = None
            if isinstance(bg_color, str):
                fill = PatternFill(
                    start_color=self._hex_no_hash(bg_color),
                    end_color=self._hex_no_hash(bg_color),
                    fill_type="solid",
                )

            rule = None
            if kind == "cellIs":
                rule = CellIsRule(
                    operator=str(op.get("operator")),
                    formula=[str(op.get("formula"))],
                    stopIfTrue=bool(op.get("stop_if_true", False)),
                    fill=fill,
                )
            elif kind == "expression":
                rule = FormulaRule(
                    formula=[str(op.get("formula"))],
                    stopIfTrue=bool(op.get("stop_if_true", False)),
                    fill=fill,
                )
            elif kind == "dataBar":
                min_obj = op.get("min_value", 0)
                max_obj = op.get("max_value", 0)
                min_value = int(min_obj) if isinstance(min_obj, (int, float, str)) else 0
                max_value = int(max_obj) if isinstance(max_obj, (int, float, str)) else 0
                rule = DataBarRule(
                    start_type="num",
                    start_value=min_value,
                    end_type="num",
                    end_value=max_value,
                    color="638EC6",
                )
            elif kind == "colorScale":
                rule = ColorScaleRule(
                    start_type="min",
                    start_color="F8696B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="63BE7B",
                )

            if rule is None:
                continue
            priority = op.get("priority")
            if isinstance(priority, int):
                rule.priority = priority
            ws.conditional_formatting.add(cf_range, rule)

        wb.save(output_path)
