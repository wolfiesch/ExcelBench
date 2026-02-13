"""Smoke/integration tests for optional Rust backends.

These tests are designed to be safe in environments where the PyO3 extension
module is not built/installed.
"""

from __future__ import annotations

import importlib.util
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pytest

from excelbench.models import BorderEdge, BorderInfo, BorderStyle, CellFormat


def _enabled_backends(excelbench_rust: Any) -> set[str]:
    info = excelbench_rust.build_info()
    if isinstance(info, dict):
        enabled = info.get("enabled_backends")
        if isinstance(enabled, list):
            return {str(x) for x in enabled}
    return set()


def test_registry_works_without_excelbench_rust() -> None:
    """If the native extension isn't installed, adapter discovery must still work."""
    if importlib.util.find_spec("excelbench_rust") is not None:
        pytest.skip("excelbench_rust installed; this test targets no-extension environments")

    from excelbench.harness.adapters import get_all_adapters

    names = [a.name for a in get_all_adapters()]
    assert "calamine" not in names
    assert "rust_xlsxwriter" not in names
    assert "umya-spreadsheet" not in names


def test_rust_calamine_datetime_semantics() -> None:
    excelbench_rust = pytest.importorskip("excelbench_rust")
    enabled = _enabled_backends(excelbench_rust)
    if "calamine" not in enabled:
        pytest.skip("excelbench_rust compiled without calamine backend")

    from excelbench.harness.adapters.openpyxl_adapter import OpenpyxlAdapter
    from excelbench.harness.adapters.rust_calamine_adapter import RustCalamineAdapter
    from excelbench.models import CellType, CellValue

    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path = Path(f.name)
    f.close()
    try:
        openpyxl = OpenpyxlAdapter()
        wb = openpyxl.create_workbook()
        openpyxl.add_sheet(wb, "S")
        openpyxl.write_cell_value(
            wb, "S", "A1", CellValue(type=CellType.DATE, value=date(2024, 6, 15))
        )
        openpyxl.write_cell_value(
            wb,
            "S",
            "A2",
            CellValue(type=CellType.DATETIME, value=datetime(2024, 6, 15, 10, 30, 0)),
        )
        openpyxl.save_workbook(wb, path)

        adapter = RustCalamineAdapter()
        wb2 = adapter.open_workbook(path)
        v1 = adapter.read_cell_value(wb2, "S", "A1")
        v2 = adapter.read_cell_value(wb2, "S", "A2")
        assert v1.type == CellType.DATE
        assert v2.type == CellType.DATETIME
    finally:
        path.unlink(missing_ok=True)


def test_rust_xlsxwriter_preserves_sheet_insertion_order() -> None:
    excelbench_rust = pytest.importorskip("excelbench_rust")
    enabled = _enabled_backends(excelbench_rust)
    if "rust_xlsxwriter" not in enabled:
        pytest.skip("excelbench_rust compiled without rust_xlsxwriter backend")

    import openpyxl

    from excelbench.harness.adapters.rust_xlsxwriter_adapter import RustXlsxWriterAdapter
    from excelbench.models import CellType, CellValue

    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path = Path(f.name)
    f.close()
    try:
        adapter = RustXlsxWriterAdapter()
        wb = adapter.create_workbook()

        # Insert in non-alphabetical order to detect ordering bugs.
        adapter.add_sheet(wb, "Sheet2")
        adapter.add_sheet(wb, "Sheet1")
        adapter.write_cell_value(wb, "Sheet2", "A1", CellValue(type=CellType.STRING, value="s2"))
        adapter.write_cell_value(wb, "Sheet1", "A1", CellValue(type=CellType.STRING, value="s1"))

        adapter.save_workbook(wb, path)

        wb2 = openpyxl.load_workbook(str(path), data_only=False)
        assert wb2.sheetnames == ["Sheet2", "Sheet1"]
        wb2.close()
    finally:
        path.unlink(missing_ok=True)


def test_rust_xlsxwriter_error_tokens_write_expected_formula() -> None:
    excelbench_rust = pytest.importorskip("excelbench_rust")
    enabled = _enabled_backends(excelbench_rust)
    if "rust_xlsxwriter" not in enabled:
        pytest.skip("excelbench_rust compiled without rust_xlsxwriter backend")

    import openpyxl

    from excelbench.harness.adapters.rust_xlsxwriter_adapter import RustXlsxWriterAdapter
    from excelbench.models import CellType, CellValue

    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path = Path(f.name)
    f.close()
    try:
        adapter = RustXlsxWriterAdapter()
        wb = adapter.create_workbook()
        adapter.add_sheet(wb, "S")
        adapter.write_cell_value(wb, "S", "A1", CellValue(type=CellType.ERROR, value="#REF!"))
        adapter.save_workbook(wb, path)

        wb2 = openpyxl.load_workbook(str(path), data_only=False)
        ws = wb2["S"]
        # For non-mapped error tokens we write the literal token as a string.
        assert ws["A1"].value == "#REF!"
        wb2.close()
    finally:
        path.unlink(missing_ok=True)


def test_rust_xlsxwriter_writes_dates_as_excel_dates() -> None:
    excelbench_rust = pytest.importorskip("excelbench_rust")
    enabled = _enabled_backends(excelbench_rust)
    if "rust_xlsxwriter" not in enabled:
        pytest.skip("excelbench_rust compiled without rust_xlsxwriter backend")

    import openpyxl

    from excelbench.harness.adapters.rust_xlsxwriter_adapter import RustXlsxWriterAdapter
    from excelbench.models import CellType, CellValue

    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path = Path(f.name)
    f.close()
    try:
        adapter = RustXlsxWriterAdapter()
        wb = adapter.create_workbook()
        adapter.add_sheet(wb, "S")
        adapter.write_cell_value(
            wb, "S", "A1", CellValue(type=CellType.DATE, value=date(2024, 6, 15))
        )
        adapter.write_cell_value(
            wb,
            "S",
            "A2",
            CellValue(type=CellType.DATETIME, value=datetime(2024, 6, 15, 10, 30, 0)),
        )
        adapter.save_workbook(wb, path)

        wb2 = openpyxl.load_workbook(str(path), data_only=False)
        ws = wb2["S"]
        assert ws["A1"].value == datetime(2024, 6, 15, 0, 0, 0)
        assert ws["A2"].value == datetime(2024, 6, 15, 10, 30, 0)
        wb2.close()
    finally:
        path.unlink(missing_ok=True)


def test_umya_write_date_datetime_and_error_encodings() -> None:
    excelbench_rust = pytest.importorskip("excelbench_rust")
    enabled = _enabled_backends(excelbench_rust)
    if "umya-spreadsheet" not in enabled:
        pytest.skip("excelbench_rust compiled without umya backend")

    import openpyxl

    from excelbench.harness.adapters.umya_adapter import UmyaAdapter
    from excelbench.models import CellType, CellValue

    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path = Path(f.name)
    f.close()
    try:
        adapter = UmyaAdapter()
        wb = adapter.create_workbook()
        adapter.add_sheet(wb, "S")
        adapter.write_cell_value(
            wb, "S", "A1", CellValue(type=CellType.DATE, value=date(2024, 6, 15))
        )
        adapter.write_cell_value(
            wb,
            "S",
            "A2",
            CellValue(type=CellType.DATETIME, value=datetime(2024, 6, 15, 10, 30, 0)),
        )
        adapter.write_cell_value(wb, "S", "A3", CellValue(type=CellType.ERROR, value="#DIV/0!"))
        adapter.write_cell_value(wb, "S", "A4", CellValue(type=CellType.ERROR, value="#REF!"))
        adapter.save_workbook(wb, path)

        wb2 = openpyxl.load_workbook(str(path), data_only=False)
        ws = wb2["S"]
        assert ws["A1"].value == datetime(2024, 6, 15, 0, 0, 0)
        assert ws["A2"].value == datetime(2024, 6, 15, 10, 30, 0)
        assert ws["A3"].data_type == "f"
        assert ws["A3"].value == "=1/0"
        assert ws["A4"].data_type == "s"
        assert ws["A4"].value == "#REF!"
        wb2.close()
    finally:
        path.unlink(missing_ok=True)


def test_umya_reads_openpyxl_dates_as_date_and_datetime() -> None:
    excelbench_rust = pytest.importorskip("excelbench_rust")
    enabled = _enabled_backends(excelbench_rust)
    if "umya-spreadsheet" not in enabled:
        pytest.skip("excelbench_rust compiled without umya backend")

    from excelbench.harness.adapters.openpyxl_adapter import OpenpyxlAdapter
    from excelbench.harness.adapters.umya_adapter import UmyaAdapter
    from excelbench.models import CellType, CellValue

    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path = Path(f.name)
    f.close()
    try:
        openpyxl = OpenpyxlAdapter()
        wb = openpyxl.create_workbook()
        openpyxl.add_sheet(wb, "S")
        openpyxl.write_cell_value(
            wb, "S", "A1", CellValue(type=CellType.DATE, value=date(2024, 6, 15))
        )
        openpyxl.write_cell_value(
            wb,
            "S",
            "A2",
            CellValue(type=CellType.DATETIME, value=datetime(2024, 6, 15, 10, 30, 0)),
        )
        openpyxl.save_workbook(wb, path)

        adapter = UmyaAdapter()
        wb2 = adapter.open_workbook(path)
        v1 = adapter.read_cell_value(wb2, "S", "A1")
        v2 = adapter.read_cell_value(wb2, "S", "A2")
        assert v1.type == CellType.DATE
        assert v2.type == CellType.DATETIME
    finally:
        path.unlink(missing_ok=True)


# ===========================================================================
# Formatting integration tests
# ===========================================================================


def _skip_unless_rust_xlsxwriter() -> None:
    excelbench_rust = pytest.importorskip("excelbench_rust")
    enabled = _enabled_backends(excelbench_rust)
    if "rust_xlsxwriter" not in enabled:
        pytest.skip("excelbench_rust compiled without rust_xlsxwriter backend")


def _skip_unless_umya() -> None:
    excelbench_rust = pytest.importorskip("excelbench_rust")
    enabled = _enabled_backends(excelbench_rust)
    if "umya-spreadsheet" not in enabled:
        pytest.skip("excelbench_rust compiled without umya backend")


def test_rust_xlsxwriter_writes_bold_italic() -> None:
    """Write bold+italic cell via rust_xlsxwriter, verify with openpyxl."""
    _skip_unless_rust_xlsxwriter()

    import openpyxl

    from excelbench.harness.adapters.rust_xlsxwriter_adapter import RustXlsxWriterAdapter
    from excelbench.models import CellType, CellValue

    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path = Path(f.name)
    f.close()
    try:
        adapter = RustXlsxWriterAdapter()
        wb = adapter.create_workbook()
        adapter.add_sheet(wb, "S")
        adapter.write_cell_value(
            wb, "S", "A1", CellValue(type=CellType.STRING, value="hello")
        )
        adapter.write_cell_format(
            wb, "S", "A1", CellFormat(bold=True, italic=True)
        )
        adapter.save_workbook(wb, path)

        wb2 = openpyxl.load_workbook(str(path))
        ws = wb2["S"]
        assert ws["A1"].font.bold is True
        assert ws["A1"].font.italic is True
        wb2.close()
    finally:
        path.unlink(missing_ok=True)


def test_rust_xlsxwriter_writes_font_and_bg_color() -> None:
    """Write font_color + bg_color via rust_xlsxwriter, verify with openpyxl."""
    _skip_unless_rust_xlsxwriter()

    import openpyxl

    from excelbench.harness.adapters.rust_xlsxwriter_adapter import RustXlsxWriterAdapter
    from excelbench.models import CellType, CellValue

    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path = Path(f.name)
    f.close()
    try:
        adapter = RustXlsxWriterAdapter()
        wb = adapter.create_workbook()
        adapter.add_sheet(wb, "S")
        adapter.write_cell_value(
            wb, "S", "A1", CellValue(type=CellType.STRING, value="colored")
        )
        adapter.write_cell_format(
            wb, "S", "A1", CellFormat(font_color="#FF0000", bg_color="#00FF00")
        )
        adapter.save_workbook(wb, path)

        wb2 = openpyxl.load_workbook(str(path))
        ws = wb2["S"]
        assert ws["A1"].font.color.rgb == "FFFF0000"
        assert ws["A1"].fill.fgColor.rgb == "FF00FF00"
        wb2.close()
    finally:
        path.unlink(missing_ok=True)


def test_rust_xlsxwriter_writes_borders() -> None:
    """Write borders via rust_xlsxwriter, verify with openpyxl."""
    _skip_unless_rust_xlsxwriter()

    import openpyxl

    from excelbench.harness.adapters.rust_xlsxwriter_adapter import RustXlsxWriterAdapter
    from excelbench.models import CellType, CellValue

    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path = Path(f.name)
    f.close()
    try:
        adapter = RustXlsxWriterAdapter()
        wb = adapter.create_workbook()
        adapter.add_sheet(wb, "S")
        adapter.write_cell_value(
            wb, "S", "A1", CellValue(type=CellType.STRING, value="bordered")
        )
        adapter.write_cell_border(
            wb,
            "S",
            "A1",
            BorderInfo(
                top=BorderEdge(style=BorderStyle.THIN, color="#FF0000"),
                bottom=BorderEdge(style=BorderStyle.MEDIUM, color="#000000"),
                left=BorderEdge(style=BorderStyle.THICK, color="#0000FF"),
                right=BorderEdge(style=BorderStyle.THIN, color="#000000"),
            ),
        )
        adapter.save_workbook(wb, path)

        wb2 = openpyxl.load_workbook(str(path))
        ws = wb2["S"]
        cell = ws["A1"]
        assert cell.border.top.style == "thin"
        assert cell.border.bottom.style == "medium"
        assert cell.border.left.style == "thick"
        assert cell.border.right.style == "thin"
        wb2.close()
    finally:
        path.unlink(missing_ok=True)


def test_rust_xlsxwriter_writes_alignment() -> None:
    """Write alignment + wrap via rust_xlsxwriter, verify with openpyxl."""
    _skip_unless_rust_xlsxwriter()

    import openpyxl

    from excelbench.harness.adapters.rust_xlsxwriter_adapter import RustXlsxWriterAdapter
    from excelbench.models import CellType, CellValue

    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path = Path(f.name)
    f.close()
    try:
        adapter = RustXlsxWriterAdapter()
        wb = adapter.create_workbook()
        adapter.add_sheet(wb, "S")
        adapter.write_cell_value(
            wb, "S", "A1", CellValue(type=CellType.STRING, value="centered")
        )
        adapter.write_cell_format(
            wb, "S", "A1", CellFormat(h_align="center", v_align="top", wrap=True)
        )
        adapter.save_workbook(wb, path)

        wb2 = openpyxl.load_workbook(str(path))
        ws = wb2["S"]
        assert ws["A1"].alignment.horizontal == "center"
        assert ws["A1"].alignment.vertical == "top"
        assert ws["A1"].alignment.wrapText is True
        wb2.close()
    finally:
        path.unlink(missing_ok=True)


def test_rust_xlsxwriter_writes_row_height_and_col_width() -> None:
    """Write row height + col width via rust_xlsxwriter, verify with openpyxl."""
    _skip_unless_rust_xlsxwriter()

    import openpyxl

    from excelbench.harness.adapters.rust_xlsxwriter_adapter import RustXlsxWriterAdapter
    from excelbench.models import CellType, CellValue

    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path = Path(f.name)
    f.close()
    try:
        adapter = RustXlsxWriterAdapter()
        wb = adapter.create_workbook()
        adapter.add_sheet(wb, "S")
        adapter.write_cell_value(
            wb, "S", "A1", CellValue(type=CellType.STRING, value="x")
        )
        adapter.set_row_height(wb, "S", 1, 30.0)
        adapter.set_column_width(wb, "S", "A", 20.0)
        adapter.save_workbook(wb, path)

        wb2 = openpyxl.load_workbook(str(path))
        ws = wb2["S"]
        assert ws.row_dimensions[1].height == pytest.approx(30.0, abs=0.5)
        assert ws.column_dimensions["A"].width == pytest.approx(20.0, abs=1.0)
        wb2.close()
    finally:
        path.unlink(missing_ok=True)


def test_umya_writes_bold_and_reads_back() -> None:
    """Write bold cell via umya, read back with umya — round-trip test."""
    _skip_unless_umya()

    from excelbench.harness.adapters.umya_adapter import UmyaAdapter
    from excelbench.models import CellType, CellValue

    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path = Path(f.name)
    f.close()
    try:
        adapter = UmyaAdapter()
        wb = adapter.create_workbook()
        adapter.add_sheet(wb, "S")
        adapter.write_cell_value(
            wb, "S", "A1", CellValue(type=CellType.STRING, value="bold")
        )
        adapter.write_cell_format(
            wb, "S", "A1", CellFormat(bold=True, italic=True, font_color="#FF0000")
        )
        adapter.save_workbook(wb, path)

        # Read back with umya.
        wb2 = adapter.open_workbook(path)
        fmt = adapter.read_cell_format(wb2, "S", "A1")
        assert fmt.bold is True
        assert fmt.italic is True
        assert fmt.font_color is not None
        assert fmt.font_color.upper() == "#FF0000"
    finally:
        path.unlink(missing_ok=True)


def test_umya_writes_borders_and_reads_back() -> None:
    """Write borders via umya, read back — round-trip test."""
    _skip_unless_umya()

    from excelbench.harness.adapters.umya_adapter import UmyaAdapter
    from excelbench.models import CellType, CellValue

    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path = Path(f.name)
    f.close()
    try:
        adapter = UmyaAdapter()
        wb = adapter.create_workbook()
        adapter.add_sheet(wb, "S")
        adapter.write_cell_value(
            wb, "S", "A1", CellValue(type=CellType.STRING, value="borders")
        )
        adapter.write_cell_border(
            wb,
            "S",
            "A1",
            BorderInfo(
                top=BorderEdge(style=BorderStyle.THIN, color="#FF0000"),
                bottom=BorderEdge(style=BorderStyle.MEDIUM, color="#000000"),
            ),
        )
        adapter.save_workbook(wb, path)

        # Read back.
        wb2 = adapter.open_workbook(path)
        border = adapter.read_cell_border(wb2, "S", "A1")
        assert border.top is not None
        assert border.top.style == BorderStyle.THIN
        assert border.bottom is not None
        assert border.bottom.style == BorderStyle.MEDIUM
    finally:
        path.unlink(missing_ok=True)


def test_umya_reads_openpyxl_formatted_cells() -> None:
    """Write formatted file with openpyxl, read with umya — cross-library test."""
    _skip_unless_umya()

    from excelbench.harness.adapters.openpyxl_adapter import OpenpyxlAdapter
    from excelbench.harness.adapters.umya_adapter import UmyaAdapter
    from excelbench.models import CellType, CellValue

    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path = Path(f.name)
    f.close()
    try:
        openpyxl_adapter = OpenpyxlAdapter()
        wb = openpyxl_adapter.create_workbook()
        openpyxl_adapter.add_sheet(wb, "S")
        openpyxl_adapter.write_cell_value(
            wb, "S", "A1", CellValue(type=CellType.STRING, value="styled")
        )
        openpyxl_adapter.write_cell_format(
            wb,
            "S",
            "A1",
            CellFormat(bold=True, font_size=14.0, h_align="center"),
        )
        openpyxl_adapter.save_workbook(wb, path)

        umya = UmyaAdapter()
        wb2 = umya.open_workbook(path)
        fmt = umya.read_cell_format(wb2, "S", "A1")
        assert fmt.bold is True
        assert fmt.font_size == pytest.approx(14.0, abs=0.1)
        assert fmt.h_align == "center"
    finally:
        path.unlink(missing_ok=True)


def test_umya_writes_row_height_and_col_width_roundtrip() -> None:
    """Write dimensions via umya, read back — round-trip test."""
    _skip_unless_umya()

    from excelbench.harness.adapters.umya_adapter import UmyaAdapter
    from excelbench.models import CellType, CellValue

    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path = Path(f.name)
    f.close()
    try:
        adapter = UmyaAdapter()
        wb = adapter.create_workbook()
        adapter.add_sheet(wb, "S")
        adapter.write_cell_value(
            wb, "S", "A1", CellValue(type=CellType.STRING, value="x")
        )
        adapter.set_row_height(wb, "S", 1, 30.0)
        adapter.set_column_width(wb, "S", "A", 20.0)
        adapter.save_workbook(wb, path)

        wb2 = adapter.open_workbook(path)
        h = adapter.read_row_height(wb2, "S", 1)
        w = adapter.read_column_width(wb2, "S", "A")
        assert h is not None
        assert h == pytest.approx(30.0, abs=0.5)
        assert w is not None
        assert w == pytest.approx(20.0, abs=0.5)
    finally:
        path.unlink(missing_ok=True)


def test_umya_write_format_verified_by_openpyxl() -> None:
    """Write formatted file with umya, verify with openpyxl."""
    _skip_unless_umya()

    import openpyxl

    from excelbench.harness.adapters.umya_adapter import UmyaAdapter
    from excelbench.models import CellType, CellValue

    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path = Path(f.name)
    f.close()
    try:
        adapter = UmyaAdapter()
        wb = adapter.create_workbook()
        adapter.add_sheet(wb, "S")
        adapter.write_cell_value(
            wb, "S", "A1", CellValue(type=CellType.STRING, value="test")
        )
        adapter.write_cell_format(
            wb,
            "S",
            "A1",
            CellFormat(bold=True, italic=True, bg_color="#FFFF00"),
        )
        adapter.write_cell_border(
            wb,
            "S",
            "A1",
            BorderInfo(
                top=BorderEdge(style=BorderStyle.THIN, color="#000000"),
                bottom=BorderEdge(style=BorderStyle.THIN, color="#000000"),
            ),
        )
        adapter.save_workbook(wb, path)

        wb2 = openpyxl.load_workbook(str(path))
        ws = wb2["S"]
        cell = ws["A1"]
        assert cell.font.bold is True
        assert cell.font.italic is True
        assert cell.border.top.style == "thin"
        assert cell.border.bottom.style == "thin"
        wb2.close()
    finally:
        path.unlink(missing_ok=True)
