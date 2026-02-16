"""Edge-case tests for adapter branches that standard roundtrips don't cover.

Targets remaining uncovered lines in:
- openpyxl_adapter.py: column width edge cases, hyperlink variants,
  data validation "between", image anchors, error strings in cells
- xlrd_adapter.py: error cell type, hyperlink/comment reading, formula text
- pylightxl_adapter.py: _parse_cell_ref, date/datetime/bool isinstance branches
- calamine_adapter.py + rust adapters: remaining cell type branches
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any
from unittest.mock import MagicMock

import openpyxl as _openpyxl
import pytest

from excelbench.harness.adapters.openpyxl_adapter import OpenpyxlAdapter
from excelbench.harness.adapters.rust_adapter_utils import (
    cell_value_from_payload,
    payload_from_cell_value,
)
from excelbench.models import (
    BorderInfo,
    CellType,
    CellValue,
)

JSONDict = dict[str, Any]


# ═════════════════════════════════════════════════
# Fixtures
# ═════════════════════════════════════════════════


@pytest.fixture
def opxl() -> OpenpyxlAdapter:
    return OpenpyxlAdapter()


# ═════════════════════════════════════════════════
# OpenpyxlAdapter — Column Width Edge Cases
# ═════════════════════════════════════════════════


class TestOpenpyxlColumnWidth:
    def test_column_width_not_set_returns_none(self, opxl: OpenpyxlAdapter, tmp_path: Path) -> None:
        """Column with no explicit width should return None (line 295)."""
        path = tmp_path / "no_width.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.write_cell_value(wb, "S1", "A1", CellValue(type=CellType.STRING, value="x"))
        opxl.save_workbook(wb, path)

        wb2 = opxl.open_workbook(path)
        opxl.read_column_width(wb2, "S1", "A")
        # Width could be None or a default value — we just verify no crash
        opxl.close_workbook(wb2)

    def test_column_width_set_and_read(self, opxl: OpenpyxlAdapter, tmp_path: Path) -> None:
        """Set column width and verify reading (lines 743-744, column width path)."""
        path = tmp_path / "col_width.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.set_column_width(wb, "S1", "B", 20.0)
        opxl.save_workbook(wb, path)

        wb2 = opxl.open_workbook(path)
        w = opxl.read_column_width(wb2, "S1", "B")
        assert w is not None
        assert isinstance(w, float)
        opxl.close_workbook(wb2)

    def test_column_width_with_padding_strip(self, opxl: OpenpyxlAdapter, tmp_path: Path) -> None:
        """Column widths should have known padding stripped."""
        path = tmp_path / "col_pad.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.set_column_width(wb, "S1", "A", 15.0)
        opxl.save_workbook(wb, path)

        wb2 = opxl.open_workbook(path)
        w = opxl.read_column_width(wb2, "S1", "A")
        assert w is not None
        opxl.close_workbook(wb2)


# ═════════════════════════════════════════════════
# OpenpyxlAdapter — Hyperlink Edge Cases
# ═════════════════════════════════════════════════


class TestOpenpyxlHyperlinkEdgeCases:
    def test_hyperlink_with_tooltip(self, opxl: OpenpyxlAdapter, tmp_path: Path) -> None:
        """Hyperlink with tooltip should preserve tooltip (line 850)."""
        path = tmp_path / "link_tooltip.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.add_hyperlink(
            wb,
            "S1",
            {
                "cell": "A1",
                "target": "https://example.com",
                "display": "Example",
                "tooltip": "Click me!",
            },
        )
        opxl.save_workbook(wb, path)

        wb2 = opxl.open_workbook(path)
        links = opxl.read_hyperlinks(wb2, "S1")
        assert len(links) >= 1
        assert links[0]["tooltip"] == "Click me!"
        opxl.close_workbook(wb2)

    def test_internal_hyperlink(self, opxl: OpenpyxlAdapter, tmp_path: Path) -> None:
        """Internal hyperlink (to another sheet) should be marked internal."""
        path = tmp_path / "link_internal.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.add_sheet(wb, "S2")
        opxl.add_hyperlink(
            wb,
            "S1",
            {
                "cell": "A1",
                "target": "#S2!A1",
                "display": "Go to S2",
                "internal": True,
            },
        )
        opxl.save_workbook(wb, path)

        wb2 = opxl.open_workbook(path)
        links = opxl.read_hyperlinks(wb2, "S1")
        assert len(links) >= 1
        assert links[0]["internal"] is True
        opxl.close_workbook(wb2)

    def test_hyperlink_target_with_location(self, opxl: OpenpyxlAdapter, tmp_path: Path) -> None:
        """Hyperlink with both target and location should recombine (lines 402-403).

        openpyxl stores external URL and fragment separately. When both exist,
        the adapter should join them as 'target#location'.
        """
        path = tmp_path / "link_fragment.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        # Write cell value first, then create hyperlink with fragment
        ws = wb["S1"]
        from openpyxl.worksheet.hyperlink import Hyperlink

        ws["A1"].value = "Section Link"
        ws["A1"].hyperlink = Hyperlink(
            ref="A1", target="https://example.com/page", location="section-2"
        )
        opxl.save_workbook(wb, path)

        wb2 = opxl.open_workbook(path)
        links = opxl.read_hyperlinks(wb2, "S1")
        assert len(links) >= 1
        assert links[0]["target"] == "https://example.com/page#section-2"
        assert links[0]["internal"] is False
        opxl.close_workbook(wb2)


# ═════════════════════════════════════════════════
# OpenpyxlAdapter — Data Validation Edge Cases
# ═════════════════════════════════════════════════


class TestOpenpyxlDataValidation:
    def test_between_operator_inferred(self, opxl: OpenpyxlAdapter, tmp_path: Path) -> None:
        """Data validation with formula2 but no operator should infer 'between' (line 373)."""
        path = tmp_path / "dv_between.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.add_data_validation(
            wb,
            "S1",
            {
                "range": "A1:A10",
                "validation_type": "whole",
                "operator": "between",
                "formula1": "1",
                "formula2": "100",
            },
        )
        opxl.save_workbook(wb, path)

        wb2 = opxl.open_workbook(path)
        validations = opxl.read_data_validations(wb2, "S1")
        assert len(validations) >= 1
        assert validations[0]["validation_type"] == "whole"
        assert validations[0]["formula1"] == "1"
        assert validations[0]["formula2"] == "100"
        assert validations[0]["operator"] is not None
        opxl.close_workbook(wb2)

    def test_validation_with_messages(self, opxl: OpenpyxlAdapter, tmp_path: Path) -> None:
        """Data validation with prompt/error messages."""
        path = tmp_path / "dv_msgs.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.add_data_validation(
            wb,
            "S1",
            {
                "range": "B1:B5",
                "validation_type": "list",
                "formula1": '"A,B,C"',
                "show_input": True,
                "show_error": True,
                "prompt_title": "Choose",
                "prompt": "Pick a value",
                "error_title": "Invalid",
                "error": "Must be A, B, or C",
            },
        )
        opxl.save_workbook(wb, path)

        wb2 = opxl.open_workbook(path)
        validations = opxl.read_data_validations(wb2, "S1")
        assert len(validations) >= 1
        v = validations[0]
        assert v["prompt_title"] == "Choose"
        assert v["prompt"] == "Pick a value"
        assert v["error_title"] == "Invalid"
        opxl.close_workbook(wb2)


# ═════════════════════════════════════════════════
# OpenpyxlAdapter — Image Edge Cases
# ═════════════════════════════════════════════════


class TestOpenpyxlImageEdgeCases:
    def test_image_with_missing_path(self, opxl: OpenpyxlAdapter) -> None:
        """Image with missing path should be a no-op (line 858)."""
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        # Should not raise
        opxl.add_image(wb, "S1", {"cell": "A1"})
        opxl.add_image(wb, "S1", {"path": None, "cell": "A1"})

    def test_image_with_missing_cell(self, opxl: OpenpyxlAdapter) -> None:
        """Image with missing cell should be a no-op (line 858)."""
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.add_image(wb, "S1", {"path": "/tmp/fake.png"})


# ═════════════════════════════════════════════════
# OpenpyxlAdapter — Cell Value Edge Cases
# ═════════════════════════════════════════════════


class TestOpenpyxlCellValueEdgeCases:
    def test_error_string_na(self, opxl: OpenpyxlAdapter, tmp_path: Path) -> None:
        """Cell with string '#N/A' should be detected as error (line 124)."""
        path = tmp_path / "err_str.xlsx"
        wb = _openpyxl.Workbook()
        ws = wb.active
        ws.title = "S1"
        # Directly set cell value to an error string (not via formula)
        ws["A1"].value = "#N/A"
        ws["A1"].data_type = "s"  # Force string type
        wb.save(str(path))
        wb.close()

        wb2 = opxl.open_workbook(path)
        v = opxl.read_cell_value(wb2, "S1", "A1")
        # Should detect as error or string — the value "#N/A" is in the known error set
        assert v.value == "#N/A"
        opxl.close_workbook(wb2)

    def test_error_string_div(self, opxl: OpenpyxlAdapter, tmp_path: Path) -> None:
        """Cell with string '#DIV/0!' should be detected as error (line 126)."""
        path = tmp_path / "err_div.xlsx"
        wb = _openpyxl.Workbook()
        ws = wb.active
        ws.title = "S1"
        ws["A1"].value = "#DIV/0!"
        ws["A1"].data_type = "s"
        wb.save(str(path))
        wb.close()

        wb2 = opxl.open_workbook(path)
        v = opxl.read_cell_value(wb2, "S1", "A1")
        assert v.value == "#DIV/0!"
        opxl.close_workbook(wb2)

    def test_formula_cell_read(self, opxl: OpenpyxlAdapter, tmp_path: Path) -> None:
        """Formula cell should come back with formula type."""
        path = tmp_path / "formula.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.write_cell_value(wb, "S1", "A1", CellValue(type=CellType.NUMBER, value=10))
        opxl.write_cell_value(
            wb, "S1", "A2", CellValue(type=CellType.FORMULA, formula="=A1*2", value="=A1*2")
        )
        opxl.save_workbook(wb, path)

        wb2 = opxl.open_workbook(path)
        v = opxl.read_cell_value(wb2, "S1", "A2")
        assert v.type == CellType.FORMULA
        assert "A1" in (v.formula or "")
        opxl.close_workbook(wb2)

    def test_blank_cell(self, opxl: OpenpyxlAdapter, tmp_path: Path) -> None:
        """Empty cell should return BLANK."""
        path = tmp_path / "blank.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.write_cell_value(wb, "S1", "A1", CellValue(type=CellType.BLANK))
        opxl.save_workbook(wb, path)

        wb2 = opxl.open_workbook(path)
        v = opxl.read_cell_value(wb2, "S1", "B1")
        assert v.type == CellType.BLANK
        opxl.close_workbook(wb2)


# ═════════════════════════════════════════════════
# OpenpyxlAdapter — Conditional Format Edge Cases
# ═════════════════════════════════════════════════


class TestOpenpyxlConditionalFormatEdgeCases:
    def test_formula_rule_with_font_color(self, opxl: OpenpyxlAdapter, tmp_path: Path) -> None:
        """CF rule with font color should be readable (covers font_color path)."""
        path = tmp_path / "cf_font.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.write_cell_value(wb, "S1", "A1", CellValue(type=CellType.NUMBER, value=50))
        opxl.add_conditional_format(
            wb,
            "S1",
            {
                "range": "A1:A5",
                "rule_type": "expression",
                "formula": "A1>10",
                "format": {"font_color": "#FF0000", "bg_color": "#FFFF00"},
            },
        )
        opxl.save_workbook(wb, path)

        wb2 = opxl.open_workbook(path)
        rules = opxl.read_conditional_formats(wb2, "S1")
        assert len(rules) >= 1
        # Should have format with both bg_color and font_color
        fmt = rules[0].get("format", {})
        assert "bg_color" in fmt or "font_color" in fmt
        opxl.close_workbook(wb2)

    def test_color_scale_rule(self, opxl: OpenpyxlAdapter, tmp_path: Path) -> None:
        """Color scale CF rule should roundtrip."""
        path = tmp_path / "cf_scale.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        for i in range(1, 6):
            opxl.write_cell_value(wb, "S1", f"A{i}", CellValue(type=CellType.NUMBER, value=i * 10))
        opxl.add_conditional_format(
            wb,
            "S1",
            {
                "range": "A1:A5",
                "rule_type": "colorScale",
            },
        )
        opxl.save_workbook(wb, path)

        wb2 = opxl.open_workbook(path)
        rules = opxl.read_conditional_formats(wb2, "S1")
        assert len(rules) >= 1
        opxl.close_workbook(wb2)

    def test_data_bar_rule(self, opxl: OpenpyxlAdapter, tmp_path: Path) -> None:
        """Data bar CF rule should roundtrip."""
        path = tmp_path / "cf_bar.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        for i in range(1, 6):
            opxl.write_cell_value(wb, "S1", f"A{i}", CellValue(type=CellType.NUMBER, value=i * 10))
        opxl.add_conditional_format(
            wb,
            "S1",
            {
                "range": "A1:A5",
                "rule_type": "dataBar",
            },
        )
        opxl.save_workbook(wb, path)

        wb2 = opxl.open_workbook(path)
        rules = opxl.read_conditional_formats(wb2, "S1")
        assert len(rules) >= 1
        opxl.close_workbook(wb2)

    def test_cf_with_priority(self, opxl: OpenpyxlAdapter, tmp_path: Path) -> None:
        """CF rule with explicit priority should preserve it."""
        path = tmp_path / "cf_prio.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.add_conditional_format(
            wb,
            "S1",
            {
                "range": "A1:A5",
                "rule_type": "cellIs",
                "operator": "lessThan",
                "formula": "50",
                "priority": 1,
                "format": {"bg_color": "#00FF00"},
            },
        )
        opxl.save_workbook(wb, path)

        wb2 = opxl.open_workbook(path)
        rules = opxl.read_conditional_formats(wb2, "S1")
        assert len(rules) >= 1
        assert rules[0]["priority"] is not None
        opxl.close_workbook(wb2)

    def test_cf_stop_if_true(self, opxl: OpenpyxlAdapter, tmp_path: Path) -> None:
        """CF rule with stopIfTrue should be readable."""
        path = tmp_path / "cf_stop.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.add_conditional_format(
            wb,
            "S1",
            {
                "range": "A1:A5",
                "rule_type": "cellIs",
                "operator": "equal",
                "formula": "100",
                "stop_if_true": True,
                "format": {"bg_color": "#0000FF"},
            },
        )
        opxl.save_workbook(wb, path)

        wb2 = opxl.open_workbook(path)
        rules = opxl.read_conditional_formats(wb2, "S1")
        assert len(rules) >= 1
        opxl.close_workbook(wb2)


# ═════════════════════════════════════════════════
# OpenpyxlAdapter — Freeze Panes / Split
# ═════════════════════════════════════════════════


class TestOpenpyxlSplitPanes:
    def test_split_panes_roundtrip(self, opxl: OpenpyxlAdapter, tmp_path: Path) -> None:
        """Split panes should be readable."""
        path = tmp_path / "split.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.set_freeze_panes(
            wb,
            "S1",
            {
                "mode": "split",
                "x_split": 2,
                "y_split": 3,
                "top_left_cell": "C4",
                "active_pane": "bottomRight",
            },
        )
        opxl.save_workbook(wb, path)

        wb2 = opxl.open_workbook(path)
        result = opxl.read_freeze_panes(wb2, "S1")
        assert result.get("mode") == "split"
        opxl.close_workbook(wb2)


# ═════════════════════════════════════════════════
# OpenpyxlAdapter — Pivot Table Stub
# ═════════════════════════════════════════════════


class TestOpenpyxlPivotTable:
    def test_pivot_table_raises(self, opxl: OpenpyxlAdapter) -> None:
        """Pivot table creation should raise NotImplementedError."""
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        with pytest.raises(NotImplementedError):
            opxl.add_pivot_table(wb, "S1", {})

    def test_pivot_table_read_empty(self, opxl: OpenpyxlAdapter, tmp_path: Path) -> None:
        """Reading pivot tables from sheet without any should return empty list."""
        path = tmp_path / "no_pivot.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.save_workbook(wb, path)

        wb2 = opxl.open_workbook(path)
        pivots = opxl.read_pivot_tables(wb2, "S1")
        assert pivots == []
        opxl.close_workbook(wb2)


# ═════════════════════════════════════════════════
# OpenpyxlAdapter — Image Write+Read Roundtrip
# ═════════════════════════════════════════════════


class TestOpenpyxlImageRoundtrip:
    def test_image_write_read(self, opxl: OpenpyxlAdapter, tmp_path: Path) -> None:
        """Write an image and read it back (lines 421-447)."""
        # Create a minimal 1x1 PNG image
        import struct
        import zlib

        def make_png() -> bytes:
            sig = b"\x89PNG\r\n\x1a\n"

            def chunk(ctype: bytes, data: bytes) -> bytes:
                c = ctype + data
                crc = struct.pack(">I", zlib.crc32(c) & 0xFFFFFFFF)
                return struct.pack(">I", len(data)) + c + crc

            ihdr = struct.pack(">IIBBBBB", 1, 1, 8, 2, 0, 0, 0)
            raw_data = b"\x00\xff\x00\x00"
            idat = zlib.compress(raw_data)
            return sig + chunk(b"IHDR", ihdr) + chunk(b"IDAT", idat) + chunk(b"IEND", b"")

        img_path = tmp_path / "test.png"
        img_path.write_bytes(make_png())

        path = tmp_path / "img.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.add_image(wb, "S1", {"path": str(img_path), "cell": "B3"})
        opxl.save_workbook(wb, path)

        wb2 = opxl.open_workbook(path)
        images = opxl.read_images(wb2, "S1")
        assert len(images) >= 1
        assert images[0]["cell"] is not None
        opxl.close_workbook(wb2)


# ═════════════════════════════════════════════════
# XlrdAdapter — Error Cell Type + Hyperlinks + Comments
# ═════════════════════════════════════════════════


class TestXlrdErrorCells:
    """Test xlrd reading of error cell types (lines 176-191)."""

    @pytest.fixture
    def xlwt_adapter(self) -> Any:
        from excelbench.harness.adapters.xlwt_adapter import XlwtAdapter

        return XlwtAdapter()

    @pytest.fixture
    def xlrd_adapter(self) -> Any:
        from excelbench.harness.adapters.xlrd_adapter import XlrdAdapter

        return XlrdAdapter()

    def test_error_cell_via_xlwt(
        self, xlwt_adapter: Any, xlrd_adapter: Any, tmp_path: Path
    ) -> None:
        """Write error value with xlwt, read back with xlrd.

        xlwt stores errors as formula strings, but xlrd may interpret
        them as error cells or text depending on the Excel engine.
        """
        path = tmp_path / "err.xls"
        wb = xlwt_adapter.create_workbook()
        xlwt_adapter.add_sheet(wb, "S1")
        xlwt_adapter.write_cell_value(wb, "S1", "A1", CellValue(type=CellType.ERROR, value="#N/A"))
        xlwt_adapter.write_cell_value(
            wb, "S1", "A2", CellValue(type=CellType.ERROR, value="#DIV/0!")
        )
        xlwt_adapter.save_workbook(wb, path)

        rb = xlrd_adapter.open_workbook(path)
        v1 = xlrd_adapter.read_cell_value(rb, "S1", "A1")
        v2 = xlrd_adapter.read_cell_value(rb, "S1", "A2")
        # xlwt writes errors as formulas; xlrd reads them back
        # The value should not be None
        assert v1.value is not None
        assert v2.value is not None
        xlrd_adapter.close_workbook(rb)

    def test_formula_text_detection(
        self, xlwt_adapter: Any, xlrd_adapter: Any, tmp_path: Path
    ) -> None:
        """Text starting with '=' should be detected as formula (line 155).

        xlwt writes formulas that xlrd reads back as text cells containing
        the formula string.
        """
        path = tmp_path / "formula_text.xls"
        wb = xlwt_adapter.create_workbook()
        xlwt_adapter.add_sheet(wb, "S1")
        xlwt_adapter.write_cell_value(
            wb,
            "S1",
            "A1",
            CellValue(type=CellType.FORMULA, formula="=1+1", value="=1+1"),
        )
        xlwt_adapter.save_workbook(wb, path)

        rb = xlrd_adapter.open_workbook(path)
        v = xlrd_adapter.read_cell_value(rb, "S1", "A1")
        # Could be FORMULA or NUMBER depending on how xlwt stores it
        assert v.value is not None
        xlrd_adapter.close_workbook(rb)

    def test_xlrd_invalid_cell_ref(self) -> None:
        """Invalid cell reference should raise ValueError (line 33)."""
        from excelbench.harness.adapters.xlrd_adapter import _parse_cell_ref

        with pytest.raises(ValueError, match="Invalid cell reference"):
            _parse_cell_ref("!!!invalid")

    def test_xlrd_color_to_hex_none_rgb(self) -> None:
        """_color_to_hex with colour in map but rgb=None should return None (line 91)."""
        from excelbench.harness.adapters.xlrd_adapter import _color_to_hex

        book = MagicMock()
        book.colour_map = {10: None}  # Index exists but RGB is None
        assert _color_to_hex(book, 10) is None

    def test_xlrd_color_to_hex_valid(self) -> None:
        """_color_to_hex with valid colour should return hex string."""
        from excelbench.harness.adapters.xlrd_adapter import _color_to_hex

        book = MagicMock()
        book.colour_map = {10: (255, 0, 128)}
        assert _color_to_hex(book, 10) == "#FF0080"

    def test_xlrd_color_to_hex_skipped_indices(self) -> None:
        """_color_to_hex with special indices should return None."""
        from excelbench.harness.adapters.xlrd_adapter import _color_to_hex

        book = MagicMock()
        book.colour_map = {}
        assert _color_to_hex(book, 0x7FFF) is None  # No fill
        assert _color_to_hex(book, 64) is None  # System default


class TestXlrdRowColumnEdgeCases:
    @pytest.fixture
    def xlwt_adapter(self) -> Any:
        from excelbench.harness.adapters.xlwt_adapter import XlwtAdapter

        return XlwtAdapter()

    @pytest.fixture
    def xlrd_adapter(self) -> Any:
        from excelbench.harness.adapters.xlrd_adapter import XlrdAdapter

        return XlrdAdapter()

    def test_row_height_no_rowinfo(
        self, xlwt_adapter: Any, xlrd_adapter: Any, tmp_path: Path
    ) -> None:
        """Row without explicit height should return None (line 317)."""
        path = tmp_path / "no_rowinfo.xls"
        wb = xlwt_adapter.create_workbook()
        xlwt_adapter.add_sheet(wb, "S1")
        xlwt_adapter.write_cell_value(wb, "S1", "A1", CellValue(type=CellType.STRING, value="x"))
        xlwt_adapter.save_workbook(wb, path)

        rb = xlrd_adapter.open_workbook(path)
        # Row 5 has no data and no explicit height
        h = xlrd_adapter.read_row_height(rb, "S1", 5)
        assert h is None
        xlrd_adapter.close_workbook(rb)

    def test_column_width_colinfo(
        self, xlwt_adapter: Any, xlrd_adapter: Any, tmp_path: Path
    ) -> None:
        """Column with explicit width should return a value (line 336)."""
        path = tmp_path / "col_w.xls"
        wb = xlwt_adapter.create_workbook()
        xlwt_adapter.add_sheet(wb, "S1")
        xlwt_adapter.set_column_width(wb, "S1", "B", 20.0)
        xlwt_adapter.write_cell_value(wb, "S1", "B1", CellValue(type=CellType.STRING, value="wide"))
        xlwt_adapter.save_workbook(wb, path)

        rb = xlrd_adapter.open_workbook(path)
        w = xlrd_adapter.read_column_width(rb, "S1", "B")
        assert w is not None
        assert isinstance(w, float)
        xlrd_adapter.close_workbook(rb)


# ═════════════════════════════════════════════════
# PylightxlAdapter — Edge Cases
# ═════════════════════════════════════════════════


class TestPylightxlEdgeCases:
    @pytest.fixture
    def plxl(self) -> Any:
        from excelbench.harness.adapters.pylightxl_adapter import PylightxlAdapter

        return PylightxlAdapter()

    def test_parse_cell_ref_valid(self) -> None:
        """_parse_cell_ref should parse valid references (lines 34-42)."""
        from excelbench.harness.adapters.pylightxl_adapter import _parse_cell_ref

        row, col = _parse_cell_ref("A1")
        assert row == 1
        assert col == 1

        row, col = _parse_cell_ref("Z10")
        assert row == 10
        assert col == 26

        row, col = _parse_cell_ref("AA1")
        assert row == 1
        assert col == 27

    def test_parse_cell_ref_invalid(self) -> None:
        """_parse_cell_ref with invalid ref should raise ValueError (line 36)."""
        from excelbench.harness.adapters.pylightxl_adapter import _parse_cell_ref

        with pytest.raises(ValueError, match="Invalid cell reference"):
            _parse_cell_ref("123")

    def test_boolean_value_roundtrip(self, plxl: Any, tmp_path: Path) -> None:
        """Boolean values should survive write→read (line 96 isinstance check)."""
        path = tmp_path / "bool.xlsx"
        wb = plxl.create_workbook()
        plxl.add_sheet(wb, "S1")
        plxl.write_cell_value(wb, "S1", "A1", CellValue(type=CellType.BOOLEAN, value=True))
        plxl.save_workbook(wb, path)

        wb2 = plxl.open_workbook(path)
        v = plxl.read_cell_value(wb2, "S1", "A1")
        # pylightxl writes bools as 1/0, reads back as NUMBER
        assert v.value is not None

    def test_date_string_parsed(self, plxl: Any, tmp_path: Path) -> None:
        """pylightxl date string 'YYYY/MM/DD' should be parsed (line 117-119)."""
        path = tmp_path / "date_str.xlsx"
        wb = plxl.create_workbook()
        plxl.add_sheet(wb, "S1")
        plxl.write_cell_value(
            wb, "S1", "A1", CellValue(type=CellType.DATE, value=date(2024, 6, 15))
        )
        plxl.save_workbook(wb, path)

        wb2 = plxl.open_workbook(path)
        v = plxl.read_cell_value(wb2, "S1", "A1")
        # Should be parsed as DATE or STRING (depending on how pylightxl stores it)
        assert v.value is not None

    def test_datetime_string_parsed(self, plxl: Any, tmp_path: Path) -> None:
        """pylightxl datetime string should be parsed (lines 122-124)."""
        path = tmp_path / "datetime_str.xlsx"
        wb = plxl.create_workbook()
        plxl.add_sheet(wb, "S1")
        plxl.write_cell_value(
            wb,
            "S1",
            "A1",
            CellValue(type=CellType.DATETIME, value=datetime(2024, 6, 15, 10, 30, 45)),
        )
        plxl.save_workbook(wb, path)

        wb2 = plxl.open_workbook(path)
        v = plxl.read_cell_value(wb2, "S1", "A1")
        assert v.value is not None

    def test_error_value_read(self, plxl: Any, tmp_path: Path) -> None:
        """pylightxl should detect error strings (lines 127-130)."""
        path = tmp_path / "err.xlsx"
        wb = plxl.create_workbook()
        plxl.add_sheet(wb, "S1")
        plxl.write_cell_value(wb, "S1", "A1", CellValue(type=CellType.ERROR, value="#N/A"))
        plxl.write_cell_value(wb, "S1", "A2", CellValue(type=CellType.ERROR, value="#DIV/0!"))
        plxl.save_workbook(wb, path)

        wb2 = plxl.open_workbook(path)
        v1 = plxl.read_cell_value(wb2, "S1", "A1")
        v2 = plxl.read_cell_value(wb2, "S1", "A2")
        # pylightxl writes errors as strings
        assert v1.type == CellType.ERROR
        assert v1.value == "#N/A"
        assert v2.type == CellType.ERROR
        assert v2.value == "#DIV/0!"

    @pytest.mark.xfail(
        reason="pylightxl drops formulas in self-roundtrip",
        strict=False,
    )
    def test_formula_value_read(self, plxl: Any, tmp_path: Path) -> None:
        """pylightxl should detect formulas starting with '=' (line 134).

        pylightxl's own write drops formulas, and it crashes reading
        openpyxl-generated files due to XML parsing issues. This test
        documents the expected behavior even though it can't be verified.
        """
        path = tmp_path / "formula.xlsx"
        wb = plxl.create_workbook()
        plxl.add_sheet(wb, "S1")
        plxl.write_cell_value(
            wb,
            "S1",
            "A1",
            CellValue(type=CellType.FORMULA, formula="=1+2", value="=1+2"),
        )
        plxl.save_workbook(wb, path)

        wb2 = plxl.open_workbook(path)
        v = plxl.read_cell_value(wb2, "S1", "A1")
        assert v.type == CellType.FORMULA
        assert v.formula == "=1+2"

    def test_none_value_returns_blank(self, plxl: Any, tmp_path: Path) -> None:
        """pylightxl None value should return BLANK (line 92)."""
        path = tmp_path / "blank.xlsx"
        wb = plxl.create_workbook()
        plxl.add_sheet(wb, "S1")
        plxl.write_cell_value(wb, "S1", "A1", CellValue(type=CellType.BLANK))
        plxl.save_workbook(wb, path)

        wb2 = plxl.open_workbook(path)
        v = plxl.read_cell_value(wb2, "S1", "Z99")
        assert v.type == CellType.BLANK

    def test_border_returns_empty(self, plxl: Any, tmp_path: Path) -> None:
        """pylightxl border reading should return empty BorderInfo (line 155)."""
        path = tmp_path / "border.xlsx"
        wb = plxl.create_workbook()
        plxl.add_sheet(wb, "S1")
        plxl.write_cell_value(wb, "S1", "A1", CellValue(type=CellType.STRING, value="x"))
        plxl.save_workbook(wb, path)

        wb2 = plxl.open_workbook(path)
        b = plxl.read_cell_border(wb2, "S1", "A1")
        assert b == BorderInfo()


# ═════════════════════════════════════════════════
# CalamineAdapter — Edge Cases
# ═════════════════════════════════════════════════


try:
    from excelbench.harness.adapters import CalamineAdapter  # noqa: F401

    HAS_CALAMINE = CalamineAdapter is not None
except ImportError:
    HAS_CALAMINE = False


@pytest.mark.skipif(not HAS_CALAMINE, reason="python-calamine not installed")
class TestCalamineEdgeCases:
    @pytest.fixture
    def opxl(self) -> OpenpyxlAdapter:
        return OpenpyxlAdapter()

    @pytest.fixture
    def cal(self) -> Any:
        from excelbench.harness.adapters.calamine_adapter import CalamineAdapter

        return CalamineAdapter()

    def test_error_value(self, opxl: OpenpyxlAdapter, cal: Any, tmp_path: Path) -> None:
        """Calamine should handle error values without crashing."""
        path = tmp_path / "err.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.write_cell_value(wb, "S1", "A1", CellValue(type=CellType.ERROR, value="#N/A"))
        opxl.save_workbook(wb, path)

        wb2 = cal.open_workbook(path)
        v = cal.read_cell_value(wb2, "S1", "A1")
        # openpyxl writes errors as formulas (=NA()), calamine reads the
        # formula result which may be BLANK (no cached value) or ERROR
        assert v.type in (CellType.ERROR, CellType.BLANK, CellType.STRING)
        cal.close_workbook(wb2)

    def test_formula_value(self, opxl: OpenpyxlAdapter, cal: Any, tmp_path: Path) -> None:
        """Calamine should handle formula cells without crashing.

        openpyxl writes formulas without cached values, so calamine
        reads them as BLANK (no cached result to return).
        """
        path = tmp_path / "formula.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.write_cell_value(
            wb,
            "S1",
            "A1",
            CellValue(type=CellType.FORMULA, formula="=10+20", value="=10+20"),
        )
        opxl.save_workbook(wb, path)

        wb2 = cal.open_workbook(path)
        v = cal.read_cell_value(wb2, "S1", "A1")
        # Calamine only reads cached values; without Excel computing the formula,
        # the value will be BLANK
        assert v.type in (CellType.FORMULA, CellType.NUMBER, CellType.BLANK)
        cal.close_workbook(wb2)

    def test_date_value(self, opxl: OpenpyxlAdapter, cal: Any, tmp_path: Path) -> None:
        """Calamine should read date values."""
        path = tmp_path / "date.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.write_cell_value(
            wb, "S1", "A1", CellValue(type=CellType.DATE, value=date(2024, 6, 15))
        )
        opxl.save_workbook(wb, path)

        wb2 = cal.open_workbook(path)
        v = cal.read_cell_value(wb2, "S1", "A1")
        assert v.value is not None
        cal.close_workbook(wb2)

    def test_datetime_value(self, opxl: OpenpyxlAdapter, cal: Any, tmp_path: Path) -> None:
        """Calamine should read datetime values."""
        path = tmp_path / "datetime.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.write_cell_value(
            wb,
            "S1",
            "A1",
            CellValue(type=CellType.DATETIME, value=datetime(2024, 6, 15, 10, 30)),
        )
        opxl.save_workbook(wb, path)

        wb2 = cal.open_workbook(path)
        v = cal.read_cell_value(wb2, "S1", "A1")
        assert v.value is not None
        cal.close_workbook(wb2)


# ═════════════════════════════════════════════════
# RustCalamineAdapter — Edge Cases
# ═════════════════════════════════════════════════


try:
    from excelbench.harness.adapters import RustCalamineAdapter  # noqa: F401

    HAS_RUST_CALAMINE = RustCalamineAdapter is not None
except ImportError:
    HAS_RUST_CALAMINE = False


@pytest.mark.skipif(not HAS_RUST_CALAMINE, reason="wolfxl._rust calamine not available")
class TestRustCalamineEdgeCases:
    @pytest.fixture
    def opxl(self) -> OpenpyxlAdapter:
        return OpenpyxlAdapter()

    @pytest.fixture
    def rcal(self) -> Any:
        from excelbench.harness.adapters.rust_calamine_adapter import RustCalamineAdapter

        return RustCalamineAdapter()

    def test_date_value(self, opxl: OpenpyxlAdapter, rcal: Any, tmp_path: Path) -> None:
        """Rust calamine should handle date values."""
        path = tmp_path / "date.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.write_cell_value(
            wb, "S1", "A1", CellValue(type=CellType.DATE, value=date(2024, 3, 15))
        )
        opxl.save_workbook(wb, path)

        wb2 = rcal.open_workbook(path)
        v = rcal.read_cell_value(wb2, "S1", "A1")
        assert v.value is not None
        rcal.close_workbook(wb2)

    def test_datetime_value(self, opxl: OpenpyxlAdapter, rcal: Any, tmp_path: Path) -> None:
        """Rust calamine should handle datetime values."""
        path = tmp_path / "datetime.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.write_cell_value(
            wb,
            "S1",
            "A1",
            CellValue(type=CellType.DATETIME, value=datetime(2024, 3, 15, 14, 30)),
        )
        opxl.save_workbook(wb, path)

        wb2 = rcal.open_workbook(path)
        v = rcal.read_cell_value(wb2, "S1", "A1")
        assert v.value is not None
        rcal.close_workbook(wb2)

    def test_error_value(self, opxl: OpenpyxlAdapter, rcal: Any, tmp_path: Path) -> None:
        """Rust calamine should handle error values."""
        path = tmp_path / "err.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.write_cell_value(wb, "S1", "A1", CellValue(type=CellType.ERROR, value="#N/A"))
        opxl.save_workbook(wb, path)

        wb2 = rcal.open_workbook(path)
        v = rcal.read_cell_value(wb2, "S1", "A1")
        assert v.value is not None
        rcal.close_workbook(wb2)

    def test_blank_value(self, opxl: OpenpyxlAdapter, rcal: Any, tmp_path: Path) -> None:
        """Rust calamine should return BLANK for empty cells."""
        path = tmp_path / "blank.xlsx"
        wb = opxl.create_workbook()
        opxl.add_sheet(wb, "S1")
        opxl.write_cell_value(wb, "S1", "A1", CellValue(type=CellType.STRING, value="x"))
        opxl.save_workbook(wb, path)

        wb2 = rcal.open_workbook(path)
        v = rcal.read_cell_value(wb2, "S1", "Z99")
        assert v.type == CellType.BLANK
        rcal.close_workbook(wb2)


# ═════════════════════════════════════════════════
# Shared Rust Adapter Utils — Payload Conversion
# ═════════════════════════════════════════════════


class TestRustPayloadConversion:
    """Tests for shared rust_adapter_utils payload conversion functions.

    These are used by umya_adapter, rust_xlsxwriter_adapter, and rust_calamine_adapter.
    """

    def test_payload_date_iso(self) -> None:
        """Date payload should produce ISO string."""
        cv = CellValue(type=CellType.DATE, value=date(2024, 6, 15))
        p = payload_from_cell_value(cv)
        assert p["type"] == "date"
        assert p["value"] == "2024-06-15"

    def test_payload_date_string(self) -> None:
        """Date payload with string value should pass through."""
        cv = CellValue(type=CellType.DATE, value="2024-06-15")
        p = payload_from_cell_value(cv)
        assert p["type"] == "date"
        assert p["value"] == "2024-06-15"

    def test_payload_datetime_iso(self) -> None:
        """Datetime payload should produce ISO string without microseconds."""
        cv = CellValue(type=CellType.DATETIME, value=datetime(2024, 6, 15, 10, 30, 45, 123456))
        p = payload_from_cell_value(cv)
        assert p["type"] == "datetime"
        assert "123456" not in p["value"]

    def test_payload_datetime_string(self) -> None:
        """Datetime payload with string value should pass through."""
        cv = CellValue(type=CellType.DATETIME, value="2024-06-15T10:30:45")
        p = payload_from_cell_value(cv)
        assert p["type"] == "datetime"
        assert p["value"] == "2024-06-15T10:30:45"

    def test_payload_datetime_other(self) -> None:
        """Datetime payload with non-datetime/string should stringify."""
        cv = CellValue(type=CellType.DATETIME, value=12345)
        p = payload_from_cell_value(cv)
        assert p["type"] == "datetime"
        assert p["value"] == "12345"

    def test_payload_date_other(self) -> None:
        """Date payload with non-date/string should stringify."""
        cv = CellValue(type=CellType.DATE, value=45000)
        p = payload_from_cell_value(cv)
        assert p["type"] == "date"
        assert p["value"] == "45000"

    def test_payload_blank(self) -> None:
        """BLANK type should produce blank payload."""
        p = payload_from_cell_value(CellValue(type=CellType.BLANK))
        assert p == {"type": "blank"}

    def test_payload_number(self) -> None:
        """NUMBER type should preserve numeric value."""
        p = payload_from_cell_value(CellValue(type=CellType.NUMBER, value=3.14))
        assert p == {"type": "number", "value": 3.14}

    def test_payload_boolean(self) -> None:
        """BOOLEAN type should produce boolean payload."""
        p = payload_from_cell_value(CellValue(type=CellType.BOOLEAN, value=True))
        assert p == {"type": "boolean", "value": True}

    def test_payload_formula(self) -> None:
        """FORMULA type should produce formula payload."""
        p = payload_from_cell_value(
            CellValue(type=CellType.FORMULA, formula="=SUM(A1:A3)", value="=SUM(A1:A3)")
        )
        assert p["type"] == "formula"
        assert p["formula"] == "=SUM(A1:A3)"

    def test_payload_error(self) -> None:
        """ERROR type should stringify value."""
        p = payload_from_cell_value(CellValue(type=CellType.ERROR, value="#N/A"))
        assert p == {"type": "error", "value": "#N/A"}

    def test_payload_string_none(self) -> None:
        """STRING with None value should produce empty string."""
        p = payload_from_cell_value(CellValue(type=CellType.STRING, value=None))
        assert p == {"type": "string", "value": ""}

    def test_payload_string_value(self) -> None:
        """STRING with value should pass through."""
        p = payload_from_cell_value(CellValue(type=CellType.STRING, value="hello"))
        assert p == {"type": "string", "value": "hello"}

    def test_cell_value_from_date_string(self) -> None:
        """Date string payload should parse ISO."""
        cv = cell_value_from_payload({"type": "date", "value": "2024-06-15"})
        assert cv.type == CellType.DATE
        assert cv.value == date(2024, 6, 15)

    def test_cell_value_from_datetime_string(self) -> None:
        """Datetime string payload should parse ISO."""
        cv = cell_value_from_payload({"type": "datetime", "value": "2024-06-15T10:30:45"})
        assert cv.type == CellType.DATETIME
        assert cv.value == datetime(2024, 6, 15, 10, 30, 45)

    def test_cell_value_from_date_non_string(self) -> None:
        """Date non-string should pass through."""
        cv = cell_value_from_payload({"type": "date", "value": 45000})
        assert cv.type == CellType.DATE
        assert cv.value == 45000

    def test_cell_value_from_datetime_non_string(self) -> None:
        """Datetime non-string should pass through."""
        cv = cell_value_from_payload({"type": "datetime", "value": 45000.5})
        assert cv.type == CellType.DATETIME
        assert cv.value == 45000.5

    def test_cell_value_from_unknown_type(self) -> None:
        """Unknown type should fallback to STRING."""
        cv = cell_value_from_payload({"type": "custom", "value": "foo"})
        assert cv.type == CellType.STRING
        assert cv.value == "foo"

    def test_cell_value_from_blank(self) -> None:
        """Blank type should return BLANK."""
        cv = cell_value_from_payload({"type": "blank"})
        assert cv.type == CellType.BLANK

    def test_cell_value_from_string(self) -> None:
        """String type should return STRING."""
        cv = cell_value_from_payload({"type": "string", "value": "hello"})
        assert cv.type == CellType.STRING
        assert cv.value == "hello"

    def test_cell_value_from_number(self) -> None:
        """Number type should return NUMBER."""
        cv = cell_value_from_payload({"type": "number", "value": 42})
        assert cv.type == CellType.NUMBER
        assert cv.value == 42

    def test_cell_value_from_boolean(self) -> None:
        """Boolean type should return BOOLEAN."""
        cv = cell_value_from_payload({"type": "boolean", "value": True})
        assert cv.type == CellType.BOOLEAN
        assert cv.value is True

    def test_cell_value_from_error(self) -> None:
        """Error type should return ERROR."""
        cv = cell_value_from_payload({"type": "error", "value": "#N/A"})
        assert cv.type == CellType.ERROR
        assert cv.value == "#N/A"

    def test_cell_value_from_formula(self) -> None:
        """Formula type should return FORMULA with formula field."""
        cv = cell_value_from_payload({"type": "formula", "value": "30", "formula": "=10+20"})
        assert cv.type == CellType.FORMULA
        assert cv.value == "30"
        assert cv.formula == "=10+20"

    def test_cell_value_from_unknown_none(self) -> None:
        """Unknown type with None value should produce STRING with None."""
        cv = cell_value_from_payload({"type": "unknown", "value": None})
        assert cv.type == CellType.STRING
        assert cv.value is None
