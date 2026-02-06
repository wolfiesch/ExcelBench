from datetime import UTC, datetime
from pathlib import Path

import xlwt
from openpyxl import Workbook

from excelbench.generator.generate import write_manifest
from excelbench.harness.adapters import XlrdAdapter
from excelbench.harness.runner import run_benchmark
from excelbench.models import Importance, Manifest
from excelbench.models import TestCase as BenchCase
from excelbench.models import TestFile as BenchFile


def _write_single_case_manifest(test_dir: Path, filename: str, file_format: str) -> None:
    manifest = Manifest(
        generated_at=datetime.now(UTC),
        excel_version="test",
        generator_version="test",
        file_format=file_format,
        files=[
            BenchFile(
                path=f"tier1/{filename}",
                feature="cell_values",
                tier=1,
                file_format=file_format,
                test_cases=[
                    BenchCase(
                        id="string_simple",
                        label="String - simple",
                        row=2,
                        expected={"type": "string", "value": "Hello"},
                        importance=Importance.BASIC,
                    )
                ],
            )
        ],
    )
    write_manifest(manifest, test_dir / "manifest.json")


def test_xlrd_xlsx_is_marked_not_applicable(tmp_path):
    test_dir = tmp_path / "xlsx_suite"
    tier_dir = test_dir / "tier1"
    tier_dir.mkdir(parents=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "cell_values"
    ws["B2"] = "Hello"
    wb.save(tier_dir / "01_cell_values.xlsx")

    _write_single_case_manifest(test_dir, "01_cell_values.xlsx", "xlsx")

    results = run_benchmark(test_dir, adapters=[XlrdAdapter()], profile="xlsx")
    assert len(results.scores) == 1
    score = results.scores[0]
    assert score.read_score is None
    assert score.write_score is None
    assert score.notes == "Not applicable: xlrd does not support .xlsx input"


def test_xlrd_xls_runs_read_tests(tmp_path):
    test_dir = tmp_path / "xls_suite"
    tier_dir = test_dir / "tier1"
    tier_dir.mkdir(parents=True)

    wb = xlwt.Workbook()
    ws = wb.add_sheet("cell_values")
    ws.write(1, 1, "Hello")
    wb.save(str(tier_dir / "01_cell_values.xls"))

    _write_single_case_manifest(test_dir, "01_cell_values.xls", "xls")

    results = run_benchmark(test_dir, adapters=[XlrdAdapter()], profile="xls")
    assert len(results.scores) == 1
    score = results.scores[0]
    assert score.read_score == 3
    assert score.write_score is None
    assert score.notes is None
