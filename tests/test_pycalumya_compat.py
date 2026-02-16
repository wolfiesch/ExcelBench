"""Tests for the pycalumya openpyxl-compatible API."""

from __future__ import annotations

from pathlib import Path

import pytest


def _require_rust() -> None:
    pytest.importorskip("excelbench_rust")


# ======================================================================
# Pure-Python unit tests (no Rust needed)
# ======================================================================


class TestUtils:
    """Coordinate conversion helpers."""

    def test_column_letter(self) -> None:
        from pycalumya._utils import column_letter

        assert column_letter(1) == "A"
        assert column_letter(26) == "Z"
        assert column_letter(27) == "AA"
        assert column_letter(702) == "ZZ"

    def test_column_index(self) -> None:
        from pycalumya._utils import column_index

        assert column_index("A") == 1
        assert column_index("Z") == 26
        assert column_index("AA") == 27
        assert column_index("ZZ") == 702

    def test_a1_roundtrip(self) -> None:
        from pycalumya._utils import a1_to_rowcol, rowcol_to_a1

        assert a1_to_rowcol("B3") == (3, 2)
        assert rowcol_to_a1(3, 2) == "B3"
        assert a1_to_rowcol("AA100") == (100, 27)
        assert rowcol_to_a1(100, 27) == "AA100"

    def test_invalid_a1_raises(self) -> None:
        from pycalumya._utils import a1_to_rowcol

        with pytest.raises(ValueError, match="Invalid A1 reference"):
            a1_to_rowcol("123")


class TestStyles:
    """Frozen style dataclasses."""

    def test_font_defaults(self) -> None:
        from pycalumya._styles import Font

        f = Font()
        assert f.bold is False
        assert f.name is None
        assert f.size is None

    def test_font_is_frozen(self) -> None:
        from pycalumya._styles import Font

        f = Font(bold=True)
        with pytest.raises(AttributeError):
            f.bold = False  # type: ignore[misc]

    def test_color_hex_conversion(self) -> None:
        from pycalumya._styles import Color

        c = Color(rgb="FFFF0000")
        assert c.to_hex() == "#FF0000"
        assert Color.from_hex("#00FF00").rgb == "FF00FF00"

    def test_pattern_fill(self) -> None:
        from pycalumya._styles import PatternFill

        fill = PatternFill(patternType="solid", fgColor="#FF0000")
        assert fill._fg_hex() == "#FF0000"

    def test_border_defaults(self) -> None:
        from pycalumya._styles import Border, Side

        b = Border()
        assert b.left == Side()
        assert b.top.style is None

    def test_alignment_defaults(self) -> None:
        from pycalumya._styles import Alignment

        a = Alignment()
        assert a.horizontal is None
        assert a.wrap_text is False
        assert a.indent == 0


# ======================================================================
# Read tests (require excelbench_rust + fixtures)
# ======================================================================

FIXTURES = Path(__file__).parent.parent / "fixtures" / "excel"


class TestReadMode:
    """Read an existing Excel fixture via CalamineStyledBook."""

    def setup_method(self) -> None:
        _require_rust()

    def test_load_workbook_basic(self) -> None:
        from pycalumya import load_workbook

        path = FIXTURES / "tier1" / "01_cell_values.xlsx"
        if not path.exists():
            pytest.skip("fixture not found")
        wb = load_workbook(str(path))
        assert "Sheet1" in wb.sheetnames or len(wb.sheetnames) > 0
        ws = wb[wb.sheetnames[0]]
        # Column B has test values, A has labels. Row 2 = "Hello World" per manifest.
        val = ws["B2"].value
        assert val == "Hello World"
        wb.close()

    def test_read_number(self) -> None:
        from pycalumya import load_workbook

        path = FIXTURES / "tier1" / "01_cell_values.xlsx"
        if not path.exists():
            pytest.skip("fixture not found")
        wb = load_workbook(str(path))
        ws = wb[wb.sheetnames[0]]
        # Row 7 col B = integer 42
        val = ws["B7"].value
        assert val == 42 or val == 42.0
        wb.close()

    def test_read_font_bold(self) -> None:
        from pycalumya import load_workbook

        path = FIXTURES / "tier1" / "03_text_formatting.xlsx"
        if not path.exists():
            pytest.skip("fixture not found")
        wb = load_workbook(str(path))
        ws = wb[wb.sheetnames[0]]
        # Row 2, col B = bold text per manifest
        cell = ws["B2"]
        assert cell.font.bold is True
        wb.close()

    def test_read_background_color(self) -> None:
        from pycalumya import load_workbook

        path = FIXTURES / "tier1" / "04_background_colors.xlsx"
        if not path.exists():
            pytest.skip("fixture not found")
        wb = load_workbook(str(path))
        ws = wb[wb.sheetnames[0]]
        # Row 2, col B should have a background color
        fill = ws["B2"].fill
        # Just verify it parsed without error — exact color varies by fixture
        assert fill is not None
        wb.close()

    def test_context_manager(self) -> None:
        from pycalumya import load_workbook

        path = FIXTURES / "tier1" / "01_cell_values.xlsx"
        if not path.exists():
            pytest.skip("fixture not found")
        with load_workbook(str(path)) as wb:
            assert len(wb.sheetnames) > 0

    def test_iter_rows_read_mode(self) -> None:
        from pycalumya import load_workbook

        path = FIXTURES / "tier1" / "01_cell_values.xlsx"
        if not path.exists():
            pytest.skip("fixture not found")
        wb = load_workbook(str(path))
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
        assert len(rows) == 3
        assert rows[1][1] == "Hello World"  # B2
        wb.close()

    def test_iter_rows_auto_dimensions(self) -> None:
        from pycalumya import load_workbook

        path = FIXTURES / "tier1" / "01_cell_values.xlsx"
        if not path.exists():
            pytest.skip("fixture not found")
        wb = load_workbook(str(path))
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) > 10  # fixture has ~20 rows
        wb.close()

    def test_workbook_contains(self) -> None:
        from pycalumya import load_workbook

        path = FIXTURES / "tier1" / "01_cell_values.xlsx"
        if not path.exists():
            pytest.skip("fixture not found")
        wb = load_workbook(str(path))
        first = wb.sheetnames[0]
        assert first in wb
        assert "NonexistentSheet" not in wb
        wb.close()


# ======================================================================
# Write tests (require excelbench_rust)
# ======================================================================


class TestWriteMode:
    """Write a new Excel file via RustXlsxWriterBook."""

    def setup_method(self) -> None:
        _require_rust()

    def test_write_basic(self, tmp_path: Path) -> None:
        from pycalumya import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Hello"
        ws["B1"] = 42
        ws["C1"] = True
        out = tmp_path / "basic.xlsx"
        wb.save(str(out))
        assert out.exists()
        assert out.stat().st_size > 0

    def test_write_with_font(self, tmp_path: Path) -> None:
        from pycalumya import Font, Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Bold"
        ws["A1"].font = Font(bold=True, size=14, name="Arial")
        out = tmp_path / "font.xlsx"
        wb.save(str(out))
        assert out.exists()

    def test_write_with_fill(self, tmp_path: Path) -> None:
        from pycalumya import PatternFill, Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Colored"
        ws["A1"].fill = PatternFill(patternType="solid", fgColor="#FF0000")
        out = tmp_path / "fill.xlsx"
        wb.save(str(out))
        assert out.exists()

    def test_write_with_border(self, tmp_path: Path) -> None:
        from pycalumya import Border, Side, Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Bordered"
        ws["A1"].border = Border(
            left=Side(style="thin", color="#000000"),
            right=Side(style="thin", color="#000000"),
            top=Side(style="thin", color="#000000"),
            bottom=Side(style="thin", color="#000000"),
        )
        out = tmp_path / "border.xlsx"
        wb.save(str(out))
        assert out.exists()

    def test_write_with_alignment(self, tmp_path: Path) -> None:
        from pycalumya import Alignment, Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Centered"
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        out = tmp_path / "align.xlsx"
        wb.save(str(out))
        assert out.exists()

    def test_write_multiple_sheets(self, tmp_path: Path) -> None:
        from pycalumya import Workbook

        wb = Workbook()
        ws1 = wb.active
        assert ws1 is not None
        ws1["A1"] = "Sheet1 data"
        ws2 = wb.create_sheet("Data")
        ws2["A1"] = "Sheet2 data"
        assert wb.sheetnames == ["Sheet", "Data"]
        out = tmp_path / "multi.xlsx"
        wb.save(str(out))
        assert out.exists()

    def test_write_number_format(self, tmp_path: Path) -> None:
        from pycalumya import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = 42000
        ws["A1"].number_format = "$#,##0"
        out = tmp_path / "numfmt.xlsx"
        wb.save(str(out))
        assert out.exists()

    def test_cell_method(self, tmp_path: Path) -> None:
        from pycalumya import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        c = ws.cell(row=1, column=1, value="Via cell()")
        assert c.value == "Via cell()"
        assert c.coordinate == "A1"
        out = tmp_path / "cell_method.xlsx"
        wb.save(str(out))
        assert out.exists()


# ======================================================================
# Round-trip tests (write with pycalumya, read back)
# ======================================================================


class TestRoundTrip:
    """Write with pycalumya, read back with pycalumya."""

    def setup_method(self) -> None:
        _require_rust()

    def test_roundtrip_values(self, tmp_path: Path) -> None:
        from pycalumya import Workbook, load_workbook

        # Write
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "text"
        ws["A2"] = 123
        ws["A3"] = 3.14
        ws["A4"] = True
        out = tmp_path / "roundtrip.xlsx"
        wb.save(str(out))

        # Read back
        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        assert ws2["A1"].value == "text"
        assert ws2["A2"].value == 123 or ws2["A2"].value == 123.0
        assert abs(ws2["A3"].value - 3.14) < 0.001
        assert ws2["A4"].value is True
        wb2.close()

    def test_roundtrip_font(self, tmp_path: Path) -> None:
        from pycalumya import Font, Workbook, load_workbook

        # Write
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Bold"
        ws["A1"].font = Font(bold=True)
        out = tmp_path / "roundtrip_font.xlsx"
        wb.save(str(out))

        # Read back
        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        assert ws2["A1"].value == "Bold"
        assert ws2["A1"].font.bold is True
        wb2.close()

    def test_roundtrip_fill(self, tmp_path: Path) -> None:
        from pycalumya import PatternFill, Workbook, load_workbook

        # Write
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Red"
        ws["A1"].fill = PatternFill(patternType="solid", fgColor="#FF0000")
        out = tmp_path / "roundtrip_fill.xlsx"
        wb.save(str(out))

        # Read back
        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        fill = ws2["A1"].fill
        assert fill.fgColor is not None
        # The color should contain FF0000 (exact format may vary)
        fg = str(fill.fgColor).upper()
        assert "FF0000" in fg
        wb2.close()

    def test_roundtrip_formula(self, tmp_path: Path) -> None:
        from pycalumya import Workbook, load_workbook

        # Write
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = 10
        ws["A2"] = 20
        ws["A3"] = "=SUM(A1:A2)"
        out = tmp_path / "roundtrip_formula.xlsx"
        wb.save(str(out))

        # Read back — formula should be preserved as string
        wb2 = load_workbook(str(out))
        ws2 = wb2[wb2.sheetnames[0]]
        val = ws2["A3"].value
        assert val is not None
        assert "SUM" in str(val).upper()
        wb2.close()


# ======================================================================
# Modify mode tests (load existing, modify, save, verify)
# ======================================================================


FIXTURE = Path("fixtures/excel/tier1/01_cell_values.xlsx")


class TestModifyMode:
    """Test the read-modify-write path via WolfXL (XlsxPatcher)."""

    def setup_method(self) -> None:
        _require_rust()
        if not FIXTURE.exists():
            pytest.skip("tier1 fixture not available")

    def test_modify_repr(self) -> None:
        from pycalumya import load_workbook

        wb = load_workbook(str(FIXTURE), modify=True)
        assert "modify" in repr(wb)
        wb.close()

    def test_modify_string_value(self, tmp_path: Path) -> None:
        from pycalumya import load_workbook

        wb = load_workbook(str(FIXTURE), modify=True)
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Modified"
        out = tmp_path / "mod_string.xlsx"
        wb.save(str(out))
        wb.close()

        # Verify with pycalumya read
        wb2 = load_workbook(str(out))
        assert wb2.active is not None
        assert wb2.active["A1"].value == "Modified"
        wb2.close()

    def test_modify_number_value(self, tmp_path: Path) -> None:
        from pycalumya import load_workbook

        wb = load_workbook(str(FIXTURE), modify=True)
        ws = wb.active
        assert ws is not None
        ws["B2"] = 99.5
        out = tmp_path / "mod_number.xlsx"
        wb.save(str(out))
        wb.close()

        wb2 = load_workbook(str(out))
        assert wb2.active is not None
        assert abs(wb2.active["B2"].value - 99.5) < 0.001
        wb2.close()

    def test_modify_boolean_value(self, tmp_path: Path) -> None:
        from pycalumya import load_workbook

        wb = load_workbook(str(FIXTURE), modify=True)
        ws = wb.active
        assert ws is not None
        ws["C3"] = True
        out = tmp_path / "mod_bool.xlsx"
        wb.save(str(out))
        wb.close()

        wb2 = load_workbook(str(out))
        assert wb2.active is not None
        assert wb2.active["C3"].value is True
        wb2.close()

    def test_modify_formula(self, tmp_path: Path) -> None:
        from pycalumya import load_workbook

        wb = load_workbook(str(FIXTURE), modify=True)
        ws = wb.active
        assert ws is not None
        ws["D4"] = "=SUM(1,2,3)"
        out = tmp_path / "mod_formula.xlsx"
        wb.save(str(out))
        wb.close()

        # Verify formula preserved (openpyxl reads with = prefix)
        import openpyxl

        wb2 = openpyxl.load_workbook(str(out))
        assert wb2.active is not None
        assert wb2.active["D4"].value == "=SUM(1,2,3)"
        wb2.close()

    def test_modify_preserves_unchanged(self, tmp_path: Path) -> None:
        """Cells not touched should remain unchanged after save."""
        from pycalumya import load_workbook

        # Read original B1
        wb_orig = load_workbook(str(FIXTURE))
        assert wb_orig.active is not None
        orig_b1 = wb_orig.active["B1"].value
        wb_orig.close()

        # Modify only A1
        wb = load_workbook(str(FIXTURE), modify=True)
        ws = wb.active
        assert ws is not None
        ws["A1"] = "Changed"
        out = tmp_path / "mod_preserve.xlsx"
        wb.save(str(out))
        wb.close()

        # B1 should still have its original value
        wb2 = load_workbook(str(out))
        assert wb2.active is not None
        assert wb2.active["B1"].value == orig_b1
        assert wb2.active["A1"].value == "Changed"
        wb2.close()

    def test_modify_read_then_write(self, tmp_path: Path) -> None:
        """Read a value, modify it, save — the classic read-modify-write cycle."""
        from pycalumya import load_workbook

        wb = load_workbook(str(FIXTURE), modify=True)
        ws = wb.active
        assert ws is not None
        original = ws["A1"].value  # read via calamine
        ws["A1"] = f"WAS: {original}"  # write via patcher
        out = tmp_path / "mod_rmw.xlsx"
        wb.save(str(out))
        wb.close()

        wb2 = load_workbook(str(out))
        assert wb2.active is not None
        assert wb2.active["A1"].value == f"WAS: {original}"
        wb2.close()

    def test_modify_insert_new_cell(self, tmp_path: Path) -> None:
        """Insert a cell at a position that didn't exist in the original."""
        from pycalumya import load_workbook

        wb = load_workbook(str(FIXTURE), modify=True)
        ws = wb.active
        assert ws is not None
        ws["Z99"] = "New cell"
        out = tmp_path / "mod_insert.xlsx"
        wb.save(str(out))
        wb.close()

        wb2 = load_workbook(str(out))
        assert wb2.active is not None
        assert wb2.active["Z99"].value == "New cell"
        wb2.close()
