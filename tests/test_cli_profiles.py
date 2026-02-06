import json
from datetime import UTC, datetime
from pathlib import Path

import xlwt
from openpyxl import Workbook

from excelbench.cli import benchmark, benchmark_profiles
from excelbench.generator.generate import write_manifest
from excelbench.models import Importance, Manifest
from excelbench.models import TestCase as BenchCase
from excelbench.models import TestFile as BenchFile


def _write_cell_values_suite(test_dir: Path, extension: str, file_format: str) -> None:
    tier_dir = test_dir / "tier1"
    tier_dir.mkdir(parents=True, exist_ok=True)
    filename = f"01_cell_values.{extension}"

    if extension == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "cell_values"
        ws["B2"] = "Hello"
        wb.save(tier_dir / filename)
    else:
        wb = xlwt.Workbook()
        ws = wb.add_sheet("cell_values")
        ws.write(1, 1, "Hello")
        wb.save(str(tier_dir / filename))

    manifest = Manifest(
        generated_at=datetime.now(UTC),
        excel_version="test",
        generator_version="test",
        file_format=file_format,
        files=[
            BenchFile(
                path=f"tier1/{filename}",
                feature="cell_values",
                tier=1,
                file_format=file_format,
                test_cases=[
                    BenchCase(
                        id="string_simple",
                        label="String - simple",
                        row=2,
                        expected={"type": "string", "value": "Hello"},
                        importance=Importance.BASIC,
                    )
                ],
            )
        ],
    )
    write_manifest(manifest, test_dir / "manifest.json")


def test_benchmark_profile_xls_sets_metadata(tmp_path):
    xls_suite = tmp_path / "xls_suite"
    output_dir = tmp_path / "out_xls"
    _write_cell_values_suite(xls_suite, extension="xls", file_format="xls")

    benchmark(
        test_dir=xls_suite,
        output_dir=output_dir,
        features=None,
        append_results=False,
        profile="xls",
    )

    with open(output_dir / "results.json") as f:
        data = json.load(f)
    assert data["metadata"]["profile"] == "xls"


def test_benchmark_profiles_writes_split_outputs_and_index(tmp_path):
    xlsx_suite = tmp_path / "xlsx_suite"
    xls_suite = tmp_path / "xls_suite"
    output_dir = tmp_path / "results"

    _write_cell_values_suite(xlsx_suite, extension="xlsx", file_format="xlsx")
    _write_cell_values_suite(xls_suite, extension="xls", file_format="xls")

    benchmark_profiles(xlsx_tests=xlsx_suite, xls_tests=xls_suite, output_dir=output_dir)

    assert (output_dir / "xlsx" / "results.json").exists()
    assert (output_dir / "xls" / "results.json").exists()
    assert (output_dir / "README.md").exists()

    with open(output_dir / "xlsx" / "results.json") as f:
        xlsx_data = json.load(f)
    with open(output_dir / "xls" / "results.json") as f:
        xls_data = json.load(f)
    assert xlsx_data["metadata"]["profile"] == "xlsx"
    assert xls_data["metadata"]["profile"] == "xls"

    readme = (output_dir / "README.md").read_text()
    assert "[xlsx profile](./xlsx/README.md)" in readme
    assert "[xls profile](./xls/README.md)" in readme
