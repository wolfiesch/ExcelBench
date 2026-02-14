from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Border, PatternFill, Side

from excelbench.generator.generate import write_manifest
from excelbench.harness.adapters.openpyxl_adapter import OpenpyxlAdapter
from excelbench.harness.adapters.pandas_adapter import PandasAdapter
from excelbench.harness.adapters.tablib_adapter import TablibAdapter
from excelbench.models import Importance, Manifest
from excelbench.models import TestCase as BenchCase
from excelbench.models import TestFile as BenchFile
from excelbench.perf.runner import run_perf


def test_perf_workload_cell_values_records_op_count(tmp_path: Path) -> None:
    suite = tmp_path / "suite"
    suite.mkdir(parents=True, exist_ok=True)

    # Create a small input workbook for the read workload.
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "S1"
    for r in range(1, 4):
        for c in range(1, 4):
            ws.cell(row=r, column=c, value=r * 10 + c)
    (suite / "tier0").mkdir(parents=True, exist_ok=True)
    wb_path = suite / "tier0" / "00_cell_values_9.xlsx"
    wb.save(wb_path)

    workload = {
        "scenario": "cell_values_9",
        "op": "cell_value",
        "sheet": "S1",
        "range": "A1:C3",
        "start": 1,
        "step": 1,
    }

    manifest = Manifest(
        generated_at=datetime.now(UTC),
        excel_version="test",
        generator_version="test",
        file_format="xlsx",
        files=[
            BenchFile(
                path="tier0/00_cell_values_9.xlsx",
                feature="cell_values_9",
                tier=0,
                file_format="xlsx",
                test_cases=[
                    BenchCase(
                        id="cell_values_9",
                        label="Throughput: 9 cells",
                        row=1,
                        expected={"workload": workload},
                        importance=Importance.BASIC,
                    )
                ],
            )
        ],
    )
    write_manifest(manifest, suite / "manifest.json")

    results = run_perf(
        suite,
        adapters=[OpenpyxlAdapter()],
        warmup=0,
        iters=1,
        breakdown=True,
    )

    row = results.results[0]
    assert row.feature == "cell_values_9"
    assert row.library == "openpyxl"
    assert row.perf["read"] is not None
    assert row.perf["write"] is not None
    assert row.perf["read"].op_count == 9
    assert row.perf["read"].op_unit == "cells"
    assert row.perf["write"].op_count == 9
    assert row.perf["write"].op_unit == "cells"


def test_perf_workload_bg_color_records_op_count(tmp_path: Path) -> None:
    suite = tmp_path / "suite"
    suite.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "S1"

    fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    for r in range(1, 3):
        for c in range(1, 3):
            cell = ws.cell(row=r, column=c, value="Color")
            cell.fill = fill

    (suite / "tier0").mkdir(parents=True, exist_ok=True)
    wb_path = suite / "tier0" / "00_bg_4.xlsx"
    wb.save(wb_path)

    workload = {
        "scenario": "bg_4",
        "op": "bg_color",
        "sheet": "S1",
        "range": "A1:B2",
        "palette": ["#FF0000"],
    }

    manifest = Manifest(
        generated_at=datetime.now(UTC),
        excel_version="test",
        generator_version="test",
        file_format="xlsx",
        files=[
            BenchFile(
                path="tier0/00_bg_4.xlsx",
                feature="bg_4",
                tier=0,
                file_format="xlsx",
                test_cases=[
                    BenchCase(
                        id="bg_4",
                        label="Throughput: bg 4 cells",
                        row=1,
                        expected={"workload": workload},
                        importance=Importance.BASIC,
                    )
                ],
            )
        ],
    )
    write_manifest(manifest, suite / "manifest.json")

    results = run_perf(
        suite,
        adapters=[OpenpyxlAdapter()],
        warmup=0,
        iters=1,
        breakdown=False,
    )

    row = results.results[0]
    assert row.perf["read"] is not None
    assert row.perf["write"] is not None
    assert row.perf["read"].op_count == 4
    assert row.perf["write"].op_count == 4


def test_perf_workload_number_format_records_op_count(tmp_path: Path) -> None:
    suite = tmp_path / "suite"
    suite.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "S1"
    for r in range(1, 3):
        for c in range(1, 3):
            cell = ws.cell(row=r, column=c, value=1.25)
            cell.number_format = "0.00%"

    (suite / "tier0").mkdir(parents=True, exist_ok=True)
    wb_path = suite / "tier0" / "00_numfmt_4.xlsx"
    wb.save(wb_path)

    workload = {
        "scenario": "numfmt_4",
        "op": "number_format",
        "sheet": "S1",
        "range": "A1:B2",
        "number_format": "0.00%",
    }

    manifest = Manifest(
        generated_at=datetime.now(UTC),
        excel_version="test",
        generator_version="test",
        file_format="xlsx",
        files=[
            BenchFile(
                path="tier0/00_numfmt_4.xlsx",
                feature="numfmt_4",
                tier=0,
                file_format="xlsx",
                test_cases=[
                    BenchCase(
                        id="numfmt_4",
                        label="Throughput: number format 4 cells",
                        row=1,
                        expected={"workload": workload},
                        importance=Importance.BASIC,
                    )
                ],
            )
        ],
    )
    write_manifest(manifest, suite / "manifest.json")

    results = run_perf(
        suite,
        adapters=[OpenpyxlAdapter()],
        warmup=0,
        iters=1,
        breakdown=False,
    )
    row = results.results[0]
    assert row.perf["read"] is not None
    assert row.perf["write"] is not None
    assert row.perf["read"].op_count == 4
    assert row.perf["write"].op_count == 4


def test_perf_workload_alignment_records_op_count(tmp_path: Path) -> None:
    suite = tmp_path / "suite"
    suite.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "S1"
    for r in range(1, 3):
        for c in range(1, 3):
            ws.cell(row=r, column=c, value="Align")

    (suite / "tier0").mkdir(parents=True, exist_ok=True)
    wb_path = suite / "tier0" / "00_align_4.xlsx"
    wb.save(wb_path)

    workload = {
        "scenario": "align_4",
        "op": "alignment",
        "sheet": "S1",
        "range": "A1:B2",
        "h_align": "center",
        "v_align": "top",
        "wrap": True,
    }

    manifest = Manifest(
        generated_at=datetime.now(UTC),
        excel_version="test",
        generator_version="test",
        file_format="xlsx",
        files=[
            BenchFile(
                path="tier0/00_align_4.xlsx",
                feature="align_4",
                tier=0,
                file_format="xlsx",
                test_cases=[
                    BenchCase(
                        id="align_4",
                        label="Throughput: alignment 4 cells",
                        row=1,
                        expected={"workload": workload},
                        importance=Importance.BASIC,
                    )
                ],
            )
        ],
    )
    write_manifest(manifest, suite / "manifest.json")

    results = run_perf(
        suite,
        adapters=[OpenpyxlAdapter()],
        warmup=0,
        iters=1,
        breakdown=False,
    )
    row = results.results[0]
    assert row.perf["read"] is not None
    assert row.perf["write"] is not None
    assert row.perf["read"].op_count == 4
    assert row.perf["write"].op_count == 4


def test_perf_workload_border_records_op_count(tmp_path: Path) -> None:
    suite = tmp_path / "suite"
    suite.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "S1"

    side = Side(style="thin", color="FF000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    for r in range(1, 3):
        for c in range(1, 3):
            cell = ws.cell(row=r, column=c, value="Border")
            cell.border = border

    (suite / "tier0").mkdir(parents=True, exist_ok=True)
    wb_path = suite / "tier0" / "00_border_4.xlsx"
    wb.save(wb_path)

    workload = {
        "scenario": "border_4",
        "op": "border",
        "sheet": "S1",
        "range": "A1:B2",
        "border_style": "thin",
        "border_color": "#000000",
    }

    manifest = Manifest(
        generated_at=datetime.now(UTC),
        excel_version="test",
        generator_version="test",
        file_format="xlsx",
        files=[
            BenchFile(
                path="tier0/00_border_4.xlsx",
                feature="border_4",
                tier=0,
                file_format="xlsx",
                test_cases=[
                    BenchCase(
                        id="border_4",
                        label="Throughput: border 4 cells",
                        row=1,
                        expected={"workload": workload},
                        importance=Importance.BASIC,
                    )
                ],
            )
        ],
    )
    write_manifest(manifest, suite / "manifest.json")

    results = run_perf(
        suite,
        adapters=[OpenpyxlAdapter()],
        warmup=0,
        iters=1,
        breakdown=False,
    )
    row = results.results[0]
    assert row.perf["read"] is not None
    assert row.perf["write"] is not None
    assert row.perf["read"].op_count == 4
    assert row.perf["write"].op_count == 4


def test_perf_workload_bulk_read_skips_write(tmp_path: Path) -> None:
    suite = tmp_path / "suite"
    suite.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "S1"
    ws["A1"] = 1
    ws["B1"] = 2
    ws["A2"] = 3
    ws["B2"] = 4

    (suite / "tier0").mkdir(parents=True, exist_ok=True)
    wb_path = suite / "tier0" / "00_bulk_4.xlsx"
    wb.save(wb_path)

    workload = {
        "scenario": "bulk_4",
        "op": "bulk_sheet_values",
        "operations": ["read"],
        "sheet": "S1",
        "range": "A1:B2",
    }

    manifest = Manifest(
        generated_at=datetime.now(UTC),
        excel_version="test",
        generator_version="test",
        file_format="xlsx",
        files=[
            BenchFile(
                path="tier0/00_bulk_4.xlsx",
                feature="bulk_4",
                tier=0,
                file_format="xlsx",
                test_cases=[
                    BenchCase(
                        id="bulk_4",
                        label="Throughput: bulk read 4 cells",
                        row=1,
                        expected={"workload": workload},
                        importance=Importance.BASIC,
                    )
                ],
            )
        ],
    )
    write_manifest(manifest, suite / "manifest.json")

    results = run_perf(
        suite,
        adapters=[OpenpyxlAdapter()],
        warmup=0,
        iters=1,
        breakdown=False,
    )

    row = results.results[0]
    assert row.perf["read"] is not None
    assert row.perf["read"].op_count == 4
    assert row.perf["write"] is None
    assert row.notes is None


def test_perf_workload_bulk_write_skips_read(tmp_path: Path) -> None:
    suite = tmp_path / "suite"
    suite.mkdir(parents=True, exist_ok=True)

    (suite / "tier0").mkdir(parents=True, exist_ok=True)
    # No input workbook needed for write-only workloads.

    workload = {
        "scenario": "bulk_write_4",
        "op": "bulk_write_grid",
        "operations": ["write"],
        "sheet": "S1",
        "range": "A1:B2",
        "start": 1,
        "step": 1,
    }

    manifest = Manifest(
        generated_at=datetime.now(UTC),
        excel_version="test",
        generator_version="test",
        file_format="xlsx",
        files=[
            BenchFile(
                path="tier0/does_not_matter.xlsx",
                feature="bulk_write_4",
                tier=0,
                file_format="xlsx",
                test_cases=[
                    BenchCase(
                        id="bulk_write_4",
                        label="Throughput: bulk write 4 cells",
                        row=1,
                        expected={"workload": workload},
                        importance=Importance.BASIC,
                    )
                ],
            )
        ],
    )
    write_manifest(manifest, suite / "manifest.json")

    results = run_perf(
        suite,
        adapters=[OpenpyxlAdapter()],
        warmup=0,
        iters=1,
        breakdown=False,
    )

    row = results.results[0]
    assert row.perf["read"] is None
    assert row.perf["write"] is not None
    assert row.perf["write"].op_count == 4
    assert row.notes is None


def test_perf_workload_bulk_write_supports_strings_and_sparse_op_count(tmp_path: Path) -> None:
    suite = tmp_path / "suite"
    suite.mkdir(parents=True, exist_ok=True)
    (suite / "tier0").mkdir(parents=True, exist_ok=True)

    workload = {
        "scenario": "bulk_write_strings_sparse",
        "op": "bulk_write_grid",
        "operations": ["write"],
        "sheet": "S1",
        "range": "A1:C3",  # 9 cells
        "value_type": "string",
        "string_prefix": "V",
        "string_length": 8,
        "start": 1,
        "step": 1,
        "sparse_every": 2,
    }

    manifest = Manifest(
        generated_at=datetime.now(UTC),
        excel_version="test",
        generator_version="test",
        file_format="xlsx",
        files=[
            BenchFile(
                path="tier0/does_not_matter.xlsx",
                feature="bulk_write_strings_sparse",
                tier=0,
                file_format="xlsx",
                test_cases=[
                    BenchCase(
                        id="bulk_write_strings_sparse",
                        label="Throughput: bulk write strings sparse",
                        row=1,
                        expected={"workload": workload},
                        importance=Importance.BASIC,
                    )
                ],
            )
        ],
    )
    write_manifest(manifest, suite / "manifest.json")

    results = run_perf(
        suite,
        adapters=[OpenpyxlAdapter()],
        warmup=0,
        iters=1,
        breakdown=False,
    )

    row = results.results[0]
    assert row.perf["read"] is None
    assert row.perf["write"] is not None
    # With sparse_every=2, we fill indices 0,2,4,6,8 => 5 cells.
    assert row.perf["write"].op_count == 5
    assert row.notes is None


def test_perf_workload_bulk_write_works_for_pandas_and_tablib(tmp_path: Path) -> None:
    suite = tmp_path / "suite"
    suite.mkdir(parents=True, exist_ok=True)
    (suite / "tier0").mkdir(parents=True, exist_ok=True)

    workload = {
        "scenario": "bulk_write_4",
        "op": "bulk_write_grid",
        "operations": ["write"],
        "sheet": "S1",
        "range": "A1:B2",
        "start": 1,
        "step": 1,
    }

    manifest = Manifest(
        generated_at=datetime.now(UTC),
        excel_version="test",
        generator_version="test",
        file_format="xlsx",
        files=[
            BenchFile(
                path="tier0/does_not_matter.xlsx",
                feature="bulk_write_4",
                tier=0,
                file_format="xlsx",
                test_cases=[
                    BenchCase(
                        id="bulk_write_4",
                        label="Throughput: bulk write 4 cells",
                        row=1,
                        expected={"workload": workload},
                        importance=Importance.BASIC,
                    )
                ],
            )
        ],
    )
    write_manifest(manifest, suite / "manifest.json")

    results = run_perf(
        suite,
        adapters=[PandasAdapter(), TablibAdapter()],
        warmup=0,
        iters=1,
        breakdown=False,
    )

    assert len(results.results) == 2
    for row in results.results:
        assert row.perf["read"] is None
        assert row.perf["write"] is not None
        assert row.perf["write"].op_count == 4


def test_perf_workload_standardizes_size_and_phase_attribution(tmp_path: Path) -> None:
    suite = tmp_path / "suite"
    suite.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "S1"
    for r in range(1, 41):
        for c in range(1, 26):
            ws.cell(row=r, column=c, value=r * 100 + c)

    (suite / "tier0").mkdir(parents=True, exist_ok=True)
    wb_path = suite / "tier0" / "00_cell_values_1k.xlsx"
    wb.save(wb_path)

    workload = {
        "scenario": "cell_values_1k",
        "op": "cell_value",
        "sheet": "S1",
        "range": "A1:Y40",
        "start": 1,
        "step": 1,
    }

    manifest = Manifest(
        generated_at=datetime.now(UTC),
        excel_version="test",
        generator_version="test",
        file_format="xlsx",
        files=[
            BenchFile(
                path="tier0/00_cell_values_1k.xlsx",
                feature="cell_values_1k",
                tier=0,
                file_format="xlsx",
                test_cases=[
                    BenchCase(
                        id="cell_values_1k",
                        label="Throughput: 1k cells",
                        row=1,
                        expected={"workload": workload},
                        importance=Importance.BASIC,
                    )
                ],
            )
        ],
    )
    write_manifest(manifest, suite / "manifest.json")

    results = run_perf(
        suite,
        adapters=[OpenpyxlAdapter()],
        warmup=0,
        iters=1,
        breakdown=False,
    )

    row = results.results[0]
    assert row.workload_size == "small"

    assert row.perf["read"] is not None
    assert row.perf["write"] is not None
    read_phase = row.perf["read"].phase_attribution_ms
    write_phase = row.perf["write"].phase_attribution_ms
    assert read_phase is not None and read_phase["parse"] > 0
    assert write_phase is not None and write_phase["write"] > 0
    assert write_phase["verify"] == 0.0
